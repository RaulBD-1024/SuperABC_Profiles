"""
Inventory Insight App Interactiva - by Raúl Bolaños Díaz - 2025
=====================================
"""

import io
import unicodedata
import numpy as np
import pandas as pd
import seaborn as sns
import plotly.express as px
import streamlit as st

import statsmodels.api as sm
from prophet import Prophet
from prophet.plot import plot_plotly, plot_components_plotly


# Para generar PDF/plots
try:
    import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_LIBS_AVAILABLE = True
except Exception:
    PDF_LIBS_AVAILABLE = False

# -------------------------------
# Utilidades
# -------------------------------

def sanitize_filename(s: str) -> str:
    # elimina acentos y caracteres especiales, deja ascii y guiones bajos
    s = str(s)
    s = unicodedata.normalize('NFKD', s)
    s = s.encode('ascii', 'ignore').decode('ascii')
    s = s.replace(' ', '_')
    allowed = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_-"
    return ''.join(c for c in s if c in allowed)

def sanitize_colnames(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [unicodedata.normalize('NFKD', str(c)).encode('ascii','ignore').decode('ascii') for c in df.columns]
    df.columns = [c.replace(' ', '_') for c in df.columns]
    return df

def minmax_normalize(s: pd.Series) -> pd.Series:
    s = s.fillna(0)
    rng = s.max() - s.min()
    if rng == 0:
        return pd.Series(np.zeros(len(s)), index=s.index)
    return (s - s.min()) / rng

@st.cache_data(show_spinner=False, ttl=3600)  # Cache por 1 hora
def read_excel_bytes(file_bytes: bytes, sheet_name=None):
    """
    Lee un archivo Excel desde bytes.
    
    Args:
        file_bytes: Bytes del archivo Excel
        sheet_name: Nombre de la hoja a leer (None para todas las hojas)
    
    Returns:
        DataFrame si sheet_name está especificado, dict si no
    """
    try:
        if sheet_name:
            # Leer hoja específica
            return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, engine='openpyxl')
        else:
            # Leer todas las hojas
            return pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine='openpyxl')
    except Exception as e:
        # Si falla con openpyxl, intentar con xlrd para archivos .xls
        try:
            return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, engine='xlrd')
        except:
            raise e

# ABC por contribución acumulada

def safe_col(df: pd.DataFrame, name: str, alt_names=None):
    """Busca una columna tolerando espacios, mayúsculas/minúsculas o nombres alternativos."""
    if alt_names is None:
        alt_names = []
    # Diccionario para búsqueda insensible a mayúsculas/minúsculas y espacios
    alt = {c.strip().lower(): c for c in df.columns}
    key = name.strip().lower()
    # Revisar nombre principal
    if key in alt:
        return df[alt[key]]
    # Revisar nombres alternativos
    for alt_name in alt_names:
        k = alt_name.strip().lower()
        if k in alt:
            return df[alt[k]]
    raise KeyError(f"No se encontró la columna requerida: {name}")


def cycle_count_freq(zone: str) -> str:
    return {'Oro':'Semanal/Mensual','Plata':'Mensual/Trimestral','Bronce':'Trimestral/Semestral'}.get(zone, 'Trimestral')


def abc_by_contribution(series: pd.Series, A_cut: float, B_cut: float) -> pd.Series:
    df_tmp = series.rename('metric').to_frame()
    df_tmp = df_tmp.sort_values('metric', ascending=False)
    total = df_tmp['metric'].sum()
    if total <= 0:
        return pd.Series('C', index=series.index)
    df_tmp['cum_contrib'] = df_tmp['metric'].cumsum() / total
    df_tmp['cls'] = np.where(df_tmp['cum_contrib'] <= A_cut, 'A', np.where(df_tmp['cum_contrib'] <= B_cut, 'B', 'C'))
    return df_tmp['cls'].reindex(series.index)

@st.cache_data(show_spinner=False, ttl=1800)  # Cache por 30 minutos
def generate_super_abc_combinations(by_item: pd.DataFrame, criterios_seleccionados: list, cortes_abc: dict, criterios_map: dict) -> pd.DataFrame:
    """
    Genera todas las combinaciones posibles de clasificaciones ABC para múltiples criterios.
    
    Args:
        by_item: DataFrame con métricas por artículo
        criterios_seleccionados: Lista de criterios seleccionados
        cortes_abc: Diccionario con los cortes A y B para cada criterio
        criterios_map: Mapeo de nombres de criterios a columnas del DataFrame
    
    Returns:
        DataFrame con las clasificaciones ABC para cada criterio y la combinación final
    """
    import itertools
    
    # Calcular clasificación ABC para cada criterio
    for criterio in criterios_seleccionados:
        col_name = criterios_map[criterio]
        A_cut = cortes_abc[criterio]['A']
        B_cut = cortes_abc[criterio]['B']
        by_item[f'ABC_{criterio}'] = abc_by_contribution(by_item[col_name], A_cut, B_cut)
    
    # Generar todas las combinaciones posibles
    abc_values = ['A', 'B', 'C']
    combinaciones = list(itertools.product(abc_values, repeat=len(criterios_seleccionados)))
    
    # Crear la clasificación combinada
    def create_combination_class(row):
        combination = ''.join([row[f'ABC_{criterio}'] for criterio in criterios_seleccionados])
        return combination
    
    by_item['Clase_SuperABC'] = by_item.apply(create_combination_class, axis=1)
    
    return by_item

# Map zone from combined class

def map_zone(clase: str) -> str:
    """
    Asigna una zona (Oro, Plata, Bronce) considerando:
    - La letra más importante está al inicio.
    - Las letras A, B, C tienen pesos 3, 2 y 1 respectivamente.
    - Se aplica un promedio ponderado según posición.
    Compatible con clases de 2, 3 o más letras.
    """
    if not isinstance(clase, str) or clase.strip() == "":
        return "Bronce"

    clase = clase.strip().upper()
    pesos_letras = {'A': 3, 'B': 2, 'C': 1}

    # Ponderar más las letras de la izquierda
    n = len(clase)
    pesos_posicionales = list(range(n, 0, -1))  # Ej: [3,2,1] para 3 letras

    total = 0
    total_peso = 0
    for letra, peso_pos in zip(clase, pesos_posicionales):
        valor = pesos_letras.get(letra, 1)
        total += valor * peso_pos
        total_peso += peso_pos

    promedio_ponderado = total / total_peso

    # Umbrales calibrados con tus datos reales
    if promedio_ponderado >= 2.45:
        return "Oro"
    elif promedio_ponderado >= 2:
        return "Plata"
    else:
        return "Bronce"

def policy_by_demand(cv: float, intermittency: float) -> str:
    if intermittency >= 0.5:
        return 'RTP-EOQ (items intermitentes)'
    if cv < 0.5:
        return 'ROP-OUL (alta estabilidad)'
    if cv < 1.0:
        return 'ROP-EOQ (variabilidad media)'
    return 'RTP-EOQ (alta variabilidad)'

# Fill rates nuevos (usar guion ASCII)

def target_fill_rate(zone: str) -> str:
    if zone == 'Oro':
        return '99-90%'
    if zone == 'Plata':
        return '89-80%'
    if zone == 'Bronce':
        return '70-80%'
    return '80-90%'

# Week floor
from datetime import datetime

def week_floor(dt: pd.Series) -> pd.Series:
    return dt.dt.to_period('W-MON').dt.start_time

# -------------------------------
# UI
# -------------------------------
st.set_page_config(page_title='Inventory Insight App', layout='wide')
st.title('📦 Inventory Insight App')

st.markdown("""
Bienvenido a **Inventory Insight App** 🚀  

Desarrollado por Raúl Bolaños Díaz - 2025
            
Esta herramienta integral permite analizar y optimizar la gestión de inventarios mediante una clasificación **Súper ABC** multi-criterio, análisis de distribución de bodega y estudio de demandas.  

### 🎯 **Funcionalidades Principales:**

#### 📊 **1. Análisis Súper ABC Multi-Criterio**
- Clasificación combinada de productos según el principio de Pareto
- Múltiples criterios simultáneos (ventas, popularidad, rotación, volumen)
- Políticas de inventario personalizadas por categoría
- Zonificación automática de bodega (Oro, Plata, Bronce)

#### 🏗️ **2. Distribución Inteligente de Bodega**
- Análisis detallado por SKU con cálculos de inventario óptimo
- Cálculo de racks, pallets y espacio necesario por categoría
- Tablas de volúmenes y distribución porcentual
- Optimización de utilización de espacio

#### 📈 **3. Perfiles de Actividad Avanzados**
- Análisis de Pareto de popularidad
- Distribución de líneas por pedido
- Perfiles de cubicaje y carga unitaria
- Análisis cruzado líneas vs volumen

#### 🔮 **4. Análisis de Demanda**
- Series históricas individuales por SKU o Familia
- Estudios de tendencia y estacionalidad
- Alternativas para series muy cortas o irregulares (tendencia suavizada y ACF [Autocorrelación])

#### 📋 **5. Análisis de Contribución**
- Contribución por categoría ABC a ventas, volumen y popularidad
- Visualizaciones comparativas
- Métricas de impacto por categoría
            
### 📦 **Registros y optimización de almacén:**
- Análisis de ubicaciones actuales vs sugeridas
- Sugerencias de reubicación basadas en actividad
- Sistema de registros de movimientos de producto y actualización de ocupación
- Optimización de espacio y racks

#### 📥 **6. Exportación Completa**
- Excel con múltiples hojas organizadas
- Gráficas integradas en Excel
- Perfiles de actividad detallados
- Reportes PDF profesionales

### 🚀 **Guía de uso:**
1. **Carga de datos** → Sube tu Excel con datos de ventas
2. **Configuración** → Define criterios y cortes ABC
3. **Análisis Súper ABC** → Clasificación automática multi-criterio
4. **Perfiles de Actividad** → Análisis detallado de patrones
5. **Distribución de Bodega** → Optimización de espacio y racks
6. **Demanda** → Análisis de demanda y estacionalidad avanzado por SKU o Familia
7. **Exportación** → Descarga resultados completos

ℹ️ Esta aplicación está pensada como apoyo para decisiones de **gestión de inventario y almacenamiento**, facilitando el análisis ABC tradicional y extendido.
""")

# -------------------------------
# Advertencia sobre formato del Excel
# -------------------------------
st.info("""
📂 **Configuración del archivo Excel requerida:**

El archivo debe contener **exactamente** las siguientes columnas (respetando los nombres, aunque la aplicación es tolerante a espacios y mayúsculas/minúsculas):

- `Num. Doc` → Número de documento / pedido (factura)
- `Artículo` → Identificador único del producto 
- `Familia` → Categoría principal del producto 
- `Unid. Vend` → Cantidad de unidades vendidas  
- `Monto venta` → Monto total de venta  
- `Cajas vend.` → Cantidad de cajas vendidas (requerido para forecasting y sugerencias de distribución de bodega)
- `Cant x caja.` → Cantidad de unidades por caja (opcional, pero recomendado para análisis de carga unitaria)
- `Volumen total (p3) o Volumen total (m3)` → Volumen total del producto. Puede estar en **pies³** o **metros³**. La unidad se selecciona en el panel lateral y se convertirá automáticamente para los cálculos internos.   
- `Fecha Doc` → Fecha del documento/pedido en formato DD/MM/AAAA. 
- `Num Cliente` → Identificador del cliente (opcional, pero recomendado para análisis de popularidad)
- `Num País` → Identificador del país (opcional, pero recomendado para análisis de popularidad)

⚠️ **Importante:** Si alguna columna no existe o tiene un nombre diferente, la aplicación no podrá procesar los datos correctamente.  
Asegúrate de seleccionar la unidad correcta en la barra lateral para que los cálculos de volumen sean consistentes.
""")

with st.sidebar:
    st.header('1) Cargar datos')
    uploaded_file = st.file_uploader('Excel de ventas/ordenes', type=['xlsx','xls'])
    sheet_name = st.text_input('Hoja (opcional)', help='Si tu Excel tiene múltiples hojas, especifica cuál usar. Si no especificas, se usará la primera.')
    unit_vol = st.selectbox('Unidad de volumen', ['pies3 (p3)','metros3 (m3)'])
    vol_factor = 35.3147 if unit_vol == 'metros3 (m3)' else 1.0

    # Permitir al usuario definir el volumen de una tarima
    default_tarima = 42.38 if unit_vol == 'pies3 (p3)' else 1.2
    vol_tarima = st.number_input(
        'Volumen de una tarima completa',
        min_value=0.01,
        value=default_tarima,
        help='Define el volumen de una tarima en la unidad seleccionada'
    )
    # Guardar en session_state para usarlo en PDF y análisis
    st.session_state['vol_tarima'] = vol_tarima

    st.header('2) Criterios ABC (elige múltiples)')
    criterios = {
        'Popularidad': 'popularidad',
        'Rotacion': 'rotacion_sem',
        'Ventas': 'ventas',
        'Volumen': 'volumen'
    }
    
    # Permitir selección múltiple de criterios
    criterios_seleccionados = st.multiselect(
        'Selecciona los criterios a aplicar (mínimo 2):',
        list(criterios.keys()),
        default=['Rotacion', 'Ventas'],
        help='Puedes seleccionar 2 o más criterios. Se generarán todas las combinaciones posibles.'
    )
    
    # Validar que se seleccionen al menos 2 criterios
    if len(criterios_seleccionados) < 2:
        st.warning('⚠️ Debes seleccionar al menos 2 criterios para continuar.')
        st.stop()
    
    # Mostrar información sobre las combinaciones que se generarán
    num_combinaciones = 3 ** len(criterios_seleccionados)  # A, B, C para cada criterio
    st.info(f"📊 Se generarán {num_combinaciones} combinaciones posibles (A, B, C para cada criterio)")
    
    # Para compatibilidad con el código existente, mantener crit1 y crit2
    crit1 = criterios_seleccionados[0]
    crit2 = criterios_seleccionados[1] if len(criterios_seleccionados) > 1 else criterios_seleccionados[0]

    st.header('3) Cortes ABC por contribucion (A, B)')
    
    # Crear sliders dinámicos para cada criterio seleccionado
    cortes_abc = {}
    for i, criterio in enumerate(criterios_seleccionados):
        st.subheader(f'Criterio: {criterio}')
        A_cut = st.slider(f'A ({criterio})', 50, 95, 80, key=f'A_cut_{criterio}_{i}') / 100.0
        B_cut = st.slider(f'B ({criterio})', int(A_cut*100)+1, 99, 95, key=f'B_cut_{criterio}_{i}') / 100.0
        cortes_abc[criterio] = {'A': A_cut, 'B': B_cut}
    
    # Guardar en session_state usando claves únicas
    st.session_state['criterios_seleccionados'] = criterios_seleccionados
    st.session_state['cortes_abc'] = cortes_abc

    # Configuración de exportación (se moverá al final)
    st.session_state['want_csv'] = st.checkbox('Permitir descarga Excel', True)
    st.session_state['gen_pdf'] = st.checkbox('Generar informe PDF', False)

if uploaded_file is None:
    st.info('Sube un Excel para comenzar')
    st.stop()

# -------------------------------
# Leer datos
# -------------------------------
try:
    df = read_excel_bytes(uploaded_file.read(), sheet_name=sheet_name or None)
    
    # Verificar si df es un diccionario (múltiples hojas)
    if isinstance(df, dict):
        st.warning("⚠️ El archivo Excel contiene múltiples hojas.")
        st.write("**Hojas disponibles:**", list(df.keys()))
        
        if sheet_name and sheet_name in df:
            df = df[sheet_name]
            st.info(f"✅ Usando la hoja especificada: '{sheet_name}'")
        else:
            # Si no se especificó hoja, usar la primera
            primera_hoja = list(df.keys())[0]
            df = df[primera_hoja]
            st.info(f"✅ Usando la primera hoja: '{primera_hoja}'")
            st.write("💡 **Tip:** Puedes especificar una hoja específica en el campo 'Hoja (opcional)' en la barra lateral")
    
    # Verificar que df es un DataFrame
    if not isinstance(df, pd.DataFrame):
        st.error("❌ Error: No se pudo cargar el archivo como DataFrame")
        st.stop()
    
    # Mostrar información del archivo cargado
    st.success(f"✅ Archivo cargado exitosamente: {uploaded_file.name}")
    st.info(f"📊 Dimensiones del archivo: {df.shape[0]} filas x {df.shape[1]} columnas")
    
    # Mostrar las columnas disponibles
    st.subheader("📋 Columnas disponibles en el archivo:")
    columnas_disponibles = list(df.columns)
    st.write(columnas_disponibles)
    
    # Verificar si existe la columna 'Artículo'
    if 'Artículo' not in df.columns:
        st.error("❌ No se encontró la columna 'Artículo' en el archivo.")
        st.write("**Columnas disponibles:**", columnas_disponibles)
        st.write("**Por favor verifica que tu archivo Excel contenga una columna llamada 'Artículo'**")
        st.stop()
    
    # Limpiar espacios y mayúsculas/minúsculas
    df['Artículo_LIMPIO'] = df['Artículo'].astype(str).str.strip().str.upper()
    
except Exception as e:
    st.error(f'Error leyendo Excel: {e}')
    st.write("**Posibles causas:**")
    st.write("- El archivo no es un Excel válido")
    st.write("- El archivo está corrupto")
    st.write("- No tienes permisos para leer el archivo")
    st.write("- El archivo está siendo usado por otra aplicación")
    st.write("- El archivo tiene múltiples hojas y no se especificó cuál usar")
    st.stop()

# map columns tolerant
try:
    st.subheader("🔍 Verificando columnas requeridas...")
    
    # Verificar cada columna requerida
    columnas_requeridas = {
        'Unid. Vend': ['Unid. Vend', 'Unidades Vendidas', 'Cantidad', 'Qty'],
        'Monto venta': ['Monto venta', 'Monto Venta', 'Valor', 'Total'],
        'Volumen total (p3)': ['Volumen total (p3)', 'Volumen total (m3)', 'Volumen total', 'Volumen'],
        'Num. Doc': ['Num. Doc', 'Num Doc', 'Documento', 'Pedido', 'Order'],
        'Fecha Doc': ['Fecha Doc', 'Fecha Doc', 'Fecha', 'Date']
    }
    
    columnas_faltantes = []
    for col_principal, alternativas in columnas_requeridas.items():
        encontrada = False
        for alt in alternativas:
            if alt in df.columns:
                encontrada = True
                break
        if not encontrada:
            columnas_faltantes.append(f"{col_principal} (alternativas: {', '.join(alternativas)})")
    
    if columnas_faltantes:
        st.error("❌ Faltan las siguientes columnas requeridas:")
        for col in columnas_faltantes:
            st.write(f"- {col}")
        st.write("**Por favor verifica que tu archivo Excel contenga todas las columnas requeridas.**")
        st.stop()
    
    # Mapear columnas
    art = df['Artículo_LIMPIO']
    unid = pd.to_numeric(safe_col(df, 'Unid. Vend'), errors='coerce').fillna(0)
    monto = pd.to_numeric(safe_col(df, 'Monto venta'), errors='coerce').fillna(0)
    vol = pd.to_numeric(safe_col(df, 'Volumen total (m3)', alt_names=['Volumen total (p3)', 'Volumen total']), errors='coerce').fillna(0) # revisar si ocupa vol_factor
    numdoc = safe_col(df, 'Num. Doc').astype(str)
    fecha = pd.to_datetime(safe_col(df, 'Fecha Doc'), errors='coerce')
    familia = safe_col(df, 'Familia').astype(str).str.strip()
    num_cliente = safe_col(df, 'Num. Cliente').astype(str)
    num_pais = safe_col(df, 'Num. País').astype(str)
    # Opcional: Unidades por caja (si existe en el Excel)
    unidades_por_caja_src = pd.to_numeric(
        safe_col(df, 'Cant x Caja', alt_names=['Cant x Caja','Cant. x Caja','Unid x Caja','Unid. x Caja','Unidades por caja','Unid por caja']),
        errors='coerce'
    )
    
    st.success("✅ Todas las columnas requeridas fueron encontradas y mapeadas correctamente")
    
except Exception as e:
    st.error(f'Error mapeando columnas: {e}')
    st.write("**Posibles causas:**")
    st.write("- Los nombres de las columnas no coinciden exactamente")
    st.write("- Hay caracteres especiales o espacios extra en los nombres")
    st.write("- El formato de los datos no es el esperado")
    st.stop()

base = pd.DataFrame({
    'Articulo': art,
    'Unidades': unid,
    'Monto': monto,
    'Volumen_m3': vol,
    'NumDoc': numdoc,
    'Fecha': fecha,
    'Familia': familia,
    'NumCliente': num_cliente,
    'NumPais': num_pais,
    'Cajas_vendidas': pd.to_numeric(safe_col(df, 'Cajas vend'), errors='coerce').fillna(0),
    'Unidades_por_caja': pd.to_numeric(safe_col(df, 'Cant x caja'), errors='coerce').fillna(0)
}).dropna(subset=['Fecha'])

# Guardar base en session_state para usarlo en PDF y perfiles
st.session_state['base'] = base

if len(base) == 0:
    st.error('No hay registros con fecha valida')
    st.stop()

st.write("Primeras filas de base:")
st.dataframe(base.head())
st.write("Suma Unidades:", base['Unidades'].sum())
st.write("Suma Cajas_vendidas:", base['Cajas_vendidas'].sum())

# -------------------------------
# Calcular Super ABC
# -------------------------------
st.subheader('▶️ Control de secciones')

if st.button('1) Calcular Súper ABC'):
    by_item = base.groupby('Articulo').agg(
        popularidad=('NumDoc', 'nunique'),
        cajas=('Cajas_vendidas', 'sum'),
        ventas=('Monto', 'sum'),
        volumen=('Volumen_m3', 'sum'),
        lineas=('NumDoc', 'count'),
        Unidades=('Unidades', 'sum')
    )

    # Rotación basada en cajas
    with st.spinner("Calculando rotación semanal..."):
        days_range = (base['Fecha'].max() - base['Fecha'].min()).days + 1
        weeks_range = max(1, days_range / 7)
        months_range = max(1, days_range / 30.44)
        years_range = max(1, days_range / 365.25)

        by_item['rotacion_sem'] = by_item['cajas'] / weeks_range
        by_item['rotacion_mes'] = by_item['cajas'] / months_range
        by_item['rotacion_anual'] = by_item['cajas'] / years_range

    # Usar la nueva función para generar combinaciones múltiples
    by_item = generate_super_abc_combinations(
        by_item,
        criterios_seleccionados,
        cortes_abc,
        criterios
    )

    # Mostrar artículos con problemas de clasificación
    # Verificar si hay valores NaN en las clasificaciones ABC
    abc_columns = [f'ABC_{criterio}' for criterio in criterios_seleccionados]
    problemas_mask = by_item[abc_columns].isna().any(axis=1) | by_item['Clase_SuperABC'].str.contains('nan')
    problemas = by_item[problemas_mask]
    
    if not problemas.empty:
        st.warning(f"Hay {len(problemas)} artículos sin clase válida. Mira la tabla abajo para revisar:")
        st.dataframe(problemas)
    else:
        st.info("Todos los artículos tienen clase válida.")

    # stats semanales
    base['WeekStart'] = week_floor(base['Fecha'])
    weekly = base.groupby(['Articulo','WeekStart']).agg(units=('Unidades','sum')).reset_index()
    with st.spinner("Calculando estadísticas semanales..."):
        stats = weekly.pivot_table(index='Articulo', values='units', aggfunc=[np.mean, np.std, lambda x: (x==0).mean()])
        stats.columns = ['mean_week','std_week','intermittency']
        by_item = by_item.join(stats, how='left')
        by_item['cv'] = by_item['std_week'] / by_item['mean_week'].replace(0, np.nan)
        by_item['cv'] = by_item['cv'].fillna(np.inf)
        by_item['intermittency'] = by_item['intermittency'].fillna(1.0)

    by_item['Zona_Bodega'] = by_item['Clase_SuperABC'].apply(map_zone)
    by_item['Política_Inv'] = [policy_by_demand(cv, ii) for cv, ii in zip(by_item['cv'], by_item['intermittency'])]
    by_item['FillRate_obj'] = by_item['Zona_Bodega'].apply(target_fill_rate)
    by_item['Frecuencia_Recuento'] = by_item['Zona_Bodega'].apply(cycle_count_freq)

    st.session_state['by_item'] = by_item
    st.session_state['criterios_seleccionados'] = criterios_seleccionados
    st.session_state['crit1_name'] = crit1
    st.session_state['crit2_name'] = crit2
    st.success(f'Súper ABC calculado correctamente con {len(criterios_seleccionados)} criterios 🎯')

    # -------------------------------
    # Guardar by_item limpio como perfil
    # -------------------------------
    export_df = by_item.reset_index().copy()
    export_df.columns = [unicodedata.normalize('NFKD', str(c)).encode('ascii','ignore').decode('ascii') for c in export_df.columns]

    if 'FillRate_obj' in export_df.columns:
        export_df['FillRate_obj'] = export_df['FillRate_obj'].astype(str).str.replace('–','-', regex=False).str.replace('—','-', regex=False)

    export_df = sanitize_colnames(export_df)
    st.session_state['perfil_by_item_sanitizado'] = export_df

    # --- Comparación de artículos únicos para detectar pérdidas ---
    articulos_excel = set(df['Artículo'].astype(str).unique())
    articulos_base = set(base['Articulo'].unique())
    articulos_by_item = set(by_item.index)

    faltan_en_base = articulos_excel - articulos_base
    faltan_en_by_item = articulos_base - articulos_by_item

    st.write(f"Total artículos en Excel: {len(articulos_excel)}")
    st.write(f"Total artículos en base (con fecha válida): {len(articulos_base)}")
    st.write(f"Total artículos en by_item (agrupados): {len(articulos_by_item)}")

    if faltan_en_base:
        st.warning(f"Artículos en Excel pero no en base (probablemente por fecha vacía o inválida): {faltan_en_base}")
    if faltan_en_by_item:
        st.warning(f"Artículos en base pero no en by_item (posible error de agrupación): {faltan_en_by_item}")
    if not faltan_en_base and not faltan_en_by_item:
        st.info("No se pierden artículos en ninguna etapa del procesamiento.")
# -------------------------------
# Mostrar resumen y perfiles
# -------------------------------
def ira_by_class(clase: str) -> str:
    """
    Determina el IRA (Inventory Record Accuracy) basado en la clase de Súper ABC.
    Para múltiples criterios, se basa en la prioridad de las letras A, B, C.
    """
    # Contar la cantidad de cada letra
    count_a = clase.count('A')
    count_b = clase.count('B')
    count_c = clase.count('C')
    
    # Determinar IRA basado en la prioridad
    if count_a >= 2:  # Múltiples A
        return '> 95%'
    elif count_a == 1 and count_b >= 1:  # Una A y al menos una B
        return '94% - 95%'
    elif count_a == 1:  # Solo una A
        return '92% - 94%'
    elif count_b >= 2:  # Múltiples B
        return '90% - 92%'
    elif count_b == 1:  # Solo una B
        return '88% - 90%'
    elif count_c >= 2:  # Múltiples C
        return '86% - 88%'
    elif count_c == 1:  # Solo una C
        return '84% - 86%'
    else:
        return '< 80%'

if 'by_item' in st.session_state:
    by_item = st.session_state['by_item']

    # -------------------------------
    # Sección 2: Perfiles de Actividad
    # -------------------------------
    with st.expander("📈 Perfiles de Actividad", expanded=False):
        # -------------------------------
        # Información de criterios
        # -------------------------------
        criterios_usados = st.session_state.get('criterios_seleccionados', [crit1, crit2])
        st.subheader(f'📋 Resumen por categoría - Criterios: {", ".join(criterios_usados)}')

        combinaciones_unicas = by_item['Clase_SuperABC'].nunique()
        st.info(f"Se generaron {combinaciones_unicas} combinaciones únicas de clasificación ABC")

        # -------------------------------
        # Rotaciones por unidades y cajas
        # -------------------------------
        days_range = (base['Fecha'].max() - base['Fecha'].min()).days + 1
        weeks_range = max(1, days_range/7)
        months_range = max(1, days_range/30)
        years_range = max(1, days_range/365)

        # Rotación unidades
        by_item['Rot_Unidades_Sem'] = by_item['Unidades'] / weeks_range
        by_item['Rot_Unidades_Mes'] = by_item['Unidades'] / months_range
        by_item['Rot_Unidades_Año'] = by_item['Unidades'] / years_range

        # Rotación cajas
        if 'Cajas_vendidas' in base.columns:
            # Sumar cajas vendidas por artículo
            by_item_cajas = base.groupby('Articulo')['Cajas_vendidas'].sum()
            # Alinear con el índice de by_item
            by_item['Cajas'] = by_item_cajas.reindex(by_item.index)
            
            # Calcular rotaciones en cajas
            by_item['Rot_Cajas_Sem'] = by_item['Cajas'] / weeks_range
            by_item['Rot_Cajas_Mes'] = by_item['Cajas'] / months_range
            by_item['Rot_Cajas_Año'] = by_item['Cajas'] / years_range
        else:
            by_item['Rot_Cajas_Sem'] = np.nan
            by_item['Rot_Cajas_Mes'] = np.nan
            by_item['Rot_Cajas_Año'] = np.nan

        # -------------------------------
        # Tabla resumen por Clase_SuperABC
        # -------------------------------
        summary = by_item.groupby('Clase_SuperABC').agg(
            Cantidad=('Clase_SuperABC','count'),
            Zona_Bodega=('Zona_Bodega','first'),
            Politica=('Política_Inv','first'),
            FillRate=('FillRate_obj','first'),
            Ventas=('ventas','sum'),
            Volumen=('volumen','sum'),
            Rot_Unidades_Sem=('Rot_Unidades_Sem','mean'),
            Rot_Unidades_Mes=('Rot_Unidades_Mes','mean'),
            Rot_Unidades_Año=('Rot_Unidades_Año','mean'),
            Rot_Cajas_Sem=('Rot_Cajas_Sem','mean'),
            Rot_Cajas_Mes=('Rot_Cajas_Mes','mean'),
            Rot_Cajas_Año=('Rot_Cajas_Año','mean'),
            Popularidad=('popularidad','mean'),
            Frecuencia_Recuento=('Frecuencia_Recuento','first')
        ).reset_index()

        # IRA
        summary['IRA'] = summary['Clase_SuperABC'].apply(ira_by_class)

        # Porcentajes de suma
        summary['% Artículos'] = (100 * summary['Cantidad'] / summary['Cantidad'].sum()).round(2)
        summary['% Ventas'] = (100 * summary['Ventas'] / summary['Ventas'].sum()).round(2)
        summary['% Volumen'] = (100 * summary['Volumen'] / summary['Volumen'].sum()).round(2)

        # Porcentajes de promedio (rotaciones y popularidad)
        for col in ['Rot_Unidades_Sem','Rot_Unidades_Mes','Rot_Unidades_Año',
                    'Rot_Cajas_Sem','Rot_Cajas_Mes','Rot_Cajas_Año','Popularidad']:
            summary[f'% {col}'] = (100 * summary[col] / summary[col].sum()).round(2)

        # -------------------------------
        # Ordenar y mostrar
        # -------------------------------
        summary = summary.sort_values('Clase_SuperABC')

        cols_display = [
            'Clase_SuperABC','Cantidad','% Artículos','Zona_Bodega','Politica','FillRate','IRA',
            'Frecuencia_Recuento','Ventas','% Ventas',
            'Volumen','% Volumen',
            'Rot_Unidades_Sem','% Rot_Unidades_Sem',
            'Rot_Unidades_Mes','% Rot_Unidades_Mes',
            'Rot_Unidades_Año','% Rot_Unidades_Año',
            'Rot_Cajas_Sem','% Rot_Cajas_Sem',
            'Rot_Cajas_Mes','% Rot_Cajas_Mes',
            'Rot_Cajas_Año','% Rot_Cajas_Año',
            'Popularidad','% Popularidad'
        ]

        summary = summary[[c for c in cols_display if c in summary.columns]]  # evita KeyError


        st.dataframe(summary)
        st.session_state['perfil_resumen'] = summary

        # Perfil: lineas por orden (distribucion %)
        st.subheader('% de órdenes por # líneas')
        lines_per_order = base.groupby('NumDoc').agg(lineas=('Articulo','nunique')).reset_index()
        dist_lines = lines_per_order.groupby('lineas').size().rename('conteo').reset_index()
        total_orders = dist_lines['conteo'].sum()
        dist_lines['%_ordenes'] = 100 * dist_lines['conteo']/ (total_orders if total_orders>0 else 1)
        st.dataframe(dist_lines.sort_values('lineas'))
        fig_lines = px.bar(dist_lines.sort_values('lineas'), x='lineas', y='%_ordenes', labels={'lineas':'Líneas por orden','%_ordenes':'% de órdenes'})
        st.plotly_chart(fig_lines, use_container_width=True)

        st.session_state['perfil_lineas'] = dist_lines

        # Perfil: cubicaje por orden
        st.subheader('% de órdenes por rango de volumen (m³)')
        cubic_per_order = base.groupby('NumDoc').agg(volumen_total=('Volumen_m3','sum')).reset_index()
        vol_bins = [-1, 0.5, 1, 2, 5, 10, 20, 50, 1e9]  # Ajusta los rangos para m³
        vol_labels = ['≤0.5', '0.5-1', '1-2', '2-5', '5-10', '10-20', '20-50', '>50']
        cubic_per_order['vol_bin'] = pd.cut(cubic_per_order['volumen_total'], bins=vol_bins, labels=vol_labels)
        dist_cubic = cubic_per_order.groupby('vol_bin').size().rename('conteo').reset_index()
        total_orders2 = dist_cubic['conteo'].sum()
        dist_cubic['%_ordenes'] = 100 * dist_cubic['conteo']/ (total_orders2 if total_orders2>0 else 1)
        st.dataframe(dist_cubic)
        fig_cubic = px.bar(dist_cubic, x='vol_bin', y='%_ordenes', labels={'vol_bin':'Rango volumen (m³)','%_ordenes':'% de órdenes'})
        st.plotly_chart(fig_cubic, use_container_width=True)

        st.session_state['perfil_cubicaje'] = dist_cubic

        # Distribucion por dia de la semana
        st.subheader('Distribución de órdenes por día de la semana')
        orders_dates = base.groupby('NumDoc').agg(fecha=('Fecha','max')).reset_index()
        orders_dates['dia'] = orders_dates['fecha'].dt.day_name()
        mapping_days = {'Monday':'Lunes','Tuesday':'Martes','Wednesday':'Miércoles','Thursday':'Jueves','Friday':'Viernes','Saturday':'Sábado','Sunday':'Domingo'}
        orders_dates['dia'] = orders_dates['dia'].replace(mapping_days)
        day_order = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
        dist_days = orders_dates.groupby('dia').size().reindex(day_order).fillna(0).astype(int).rename('conteo').reset_index()
        dist_days['%_ordenes'] = 100 * dist_days['conteo'] / (dist_days['conteo'].sum() if dist_days['conteo'].sum()>0 else 1)
        st.dataframe(dist_days)
        fig_days = px.bar(dist_days, x='dia', y='%_ordenes', labels={'dia':'Día','%_ordenes':'% de órdenes'})
        st.plotly_chart(fig_days, use_container_width=True)

        st.session_state['perfil_dias'] = dist_days

        # Preparar datos
        lv = base.groupby('NumDoc').agg(
            lineas=('Articulo','nunique'),
            volumen_total=('Volumen_m3','sum')
        ).reset_index()

        # Parámetro: volumen de una tarima completa (ajusta según tu operación)
        VOLUMEN_TARIMA = st.session_state.get('vol_tarima', 42.38)

        # % de carga unitaria respecto a una tarima
        lv['%_carga_unidad'] = 100 * lv['volumen_total'] / VOLUMEN_TARIMA
        lv['%_carga_unidad'] = lv['%_carga_unidad'].clip(upper=100)  # máximo 100%

        # Bins para % de carga unitaria
        carga_bins = list(range(0, 105, 5))
        carga_labels = [f'{i}-{i+5}%' for i in range(0, 100, 5)]
        lv['r_carga'] = pd.cut(lv['%_carga_unidad'], bins=carga_bins, labels=carga_labels, right=True, include_lowest=True)
        
        # Distribución cruzada: % líneas de pedido vs % carga unitaria
        dist_incremento = lv.groupby(['r_carga']).agg(
            pedidos=('NumDoc', 'count'),
            lineas_prom=('lineas', 'mean')
        ).reset_index()
        dist_incremento['%_lineas_pedido'] = 100 * dist_incremento['pedidos'] / dist_incremento['pedidos'].sum()

        st.subheader('Distribución por incremento de pedidos (% carga unitaria vs % de líneas de pedido)')
        st.dataframe(dist_incremento.rename(columns={'%_lineas_pedido': '% de líneas de pedido'}))
        fig_incremento = px.bar(
            dist_incremento,
            x='r_carga',
            y='%_lineas_pedido',
            labels={'r_carga': '% de carga unitaria (tarima)', '%_lineas_pedido': '% de líneas de pedido'},
            title='% de líneas de pedido por % de carga unitaria'
        )
        st.plotly_chart(fig_incremento, use_container_width=True)

        st.session_state['perfil_carga'] = dist_incremento

        # -------------------------------
        # Tabla cruzada líneas x volumen por pedido
        # -------------------------------
        st.subheader('Tabla cruzada: Líneas por pedido vs pies³ por pedido')

        # Categorías
        line_labels = ['1','2-5','6-9','10+']
        lv['r_lineas'] = pd.cut(lv['lineas'], bins=[0,1,6,10,1e9], labels=line_labels, right=True, include_lowest=True)

        vol_labels = ['≤0.5', '0.5-1', '1-2', '2-5', '5-10', '10-20', '20-50', '>50']
        lv['r_vol'] = pd.cut(lv['volumen_total'], bins=[-1, 0.5, 1, 2, 5, 10, 20, 50, 1e9], labels=vol_labels, right=True, include_lowest=True)

        # Desglose de pedidos por rango de líneas (incluyendo volumen)
        st.subheader('Desglose de pedidos por rango de líneas')
        for rango in line_labels:
            pedidos_rango = lv[lv['r_lineas'] == rango][['NumDoc', 'lineas', 'volumen_total', 'r_vol']]
            st.markdown(f"**Rango {rango}: {len(pedidos_rango)} pedidos**")
            st.dataframe(pedidos_rango.reset_index(drop=True))

        # Conteo de pedidos por línea y volumen
        ct_counts = pd.crosstab(lv['r_lineas'], lv['r_vol'], dropna=False)

        # Totales de línea y % pedidos por línea
        ct_counts['Totales'] = ct_counts.sum(axis=1)
        ct_counts['% pedidos'] = (ct_counts['Totales'] / ct_counts['Totales'].sum() * 100).round(2)

        # 🔹 Total de líneas (sumando líneas, no volumen)
        pivot_lines = pd.pivot_table(
            lv,
            index='r_lineas',
            values='lineas',
            aggfunc='sum',
            fill_value=0
        )
        total_lines_global = pivot_lines['lineas'].sum()
        pivot_lines['% linea'] = (pivot_lines['lineas'] / (total_lines_global if total_lines_global>0 else 1) * 100).round(2)

        # Crear tabla final con columnas en orden deseado
        table_final = pd.DataFrame(index=line_labels, columns=vol_labels + ['Totales','% pedidos','Total_Linea','% linea'])
        table_final[vol_labels] = ct_counts[vol_labels]
        table_final['Totales'] = ct_counts['Totales']
        table_final['% pedidos'] = ct_counts['% pedidos']
        table_final['Total_Linea'] = pivot_lines['lineas']
        table_final['% linea'] = pivot_lines['% linea']

        # Fila Totales
        totales_row = table_final[vol_labels].sum()
        totales_row['Totales'] = table_final['Totales'].sum()
        totales_row['% pedidos'] = 100
        totales_row['Total_Linea'] = table_final['Total_Linea'].sum()
        totales_row['% linea'] = 100
        table_final.loc['Totales'] = totales_row

        # ✅ Fila % pedidos (por columna, ahora bien calculada)
        total_pedidos_global = table_final.loc[line_labels, vol_labels].values.sum()
        pct_pedidos_row = (table_final.loc[line_labels, vol_labels].sum() / total_pedidos_global * 100).round(2)

        pct_pedidos_row['Totales'] = 100
        pct_pedidos_row['% pedidos'] = np.nan
        pct_pedidos_row['Total_Linea'] = np.nan
        pct_pedidos_row['% linea'] = np.nan
        table_final.loc['% pedidos'] = pct_pedidos_row

        # Fila Espacio total (volumen)
        espacio_total_row = lv.groupby('r_vol')['volumen_total'].sum()
        espacio_total_row = espacio_total_row.reindex(vol_labels, fill_value=0)
        espacio_total_row['Totales'] = espacio_total_row.sum()
        espacio_total_row['% pedidos'] = np.nan
        espacio_total_row['Total_Linea'] = np.nan       # No calcular para espacio total
        espacio_total_row['% linea'] = np.nan           # No calcular para espacio total
        table_final.loc['Espacio total'] = espacio_total_row

        # Renombrar índice
        table_final.index.name = 'Líneas por orden / Volumen por orden'

        # Mostrar tabla
        st.dataframe(table_final.round(2))

        import plotly.express as px
        import plotly.graph_objects as go

        # -------------------------------
        # 1️⃣ Barras apiladas interactivas (líneas por volumen)
        # -------------------------------
        table_plot = table_final.loc[line_labels, vol_labels].astype(float)
        table_plot_reset = table_plot.reset_index().rename(columns={'index':'Rango de líneas'})

        fig_barras = px.bar(
            table_plot_reset,
            x='Líneas por orden / Volumen por orden',
            y=vol_labels,
            labels={'value':'Cantidad de pedidos','Rango de líneas':'Líneas por orden'},
            title='Distribución de pedidos por líneas y volumen (interactivo)',
            text_auto=True
        )
        st.plotly_chart(fig_barras, use_container_width=True)

        # -------------------------------
        # 2️⃣ Gráfico de pastel interactivo (por volumen total)
        # -------------------------------
        pie_data = table_final.loc['Totales', vol_labels].astype(float)
        fig_pastel = px.pie(
            names=pie_data.index,
            values=pie_data.values,
            title='Distribución de pedidos por volumen total',
            hole=0.3  # donut
        )
        st.plotly_chart(fig_pastel, use_container_width=True)

        # -------------------------------
        # 3️⃣ Heatmap interactivo (líneas x volumen)
        # -------------------------------
        heatmap_data = table_final.loc[line_labels, vol_labels].astype(float)

        fig_heatmap = go.Figure(
            data=go.Heatmap(
                z=heatmap_data.values,
                x=heatmap_data.columns,
                y=heatmap_data.index,
                colorscale='YlGnBu',
                text=heatmap_data.values,
                texttemplate="%{text}",
                colorbar=dict(title="Cantidad de pedidos")
            )
        )
        fig_heatmap.update_layout(
            title='Heatmap: Pedidos por líneas y volumen',
            xaxis_title='Volumen por orden (m³)',
            yaxis_title='Líneas por orden'
        )
        st.plotly_chart(fig_heatmap, use_container_width=True)


        st.session_state['perfil_cruzado'] = table_final

        # Pareto popularidad
        st.subheader('Pareto de popularidad de ítems (picks acumulados)')
        pareto = by_item.sort_values('popularidad', ascending=False)[['popularidad']].copy()
        pareto['cum_picks'] = pareto['popularidad'].cumsum()
        total_picks = pareto['popularidad'].sum()
        pareto['cum_pct_picks'] = 100 * pareto['cum_picks'] / (total_picks if total_picks>0 else 1)
        pareto['sku_rank'] = np.arange(1, len(pareto)+1)
        pareto['pct_sku'] = 100 * pareto['sku_rank'] / len(pareto)
        st.dataframe(pareto.head(20))
        fig_pareto = px.line(pareto, x='pct_sku', y='cum_pct_picks', labels={'pct_sku':'% de SKU (acumulado)','cum_pct_picks':'% de picks (acumulado)'}, title='Curva de Pareto – Popularidad')
        st.plotly_chart(fig_pareto, use_container_width=True)

        st.session_state['perfil_pareto'] = pareto


    # -------------------------------
    # Datos generales
    # -------------------------------
    file_name = st.session_state.get('file_name', uploaded_file.name if uploaded_file else 'Archivo no registrado')
    sheet_used = st.session_state.get('sheet_name', sheet_name or 'Hoja no registrada')
    vol_units = st.session_state.get('vol_units', unit_vol)
    criterios_usados = st.session_state.get('criterios_seleccionados', [crit1, crit2])
    cortes_abc = st.session_state.get('cortes_abc', {})
    
    # Para compatibilidad con código existente
    crit1 = criterios_usados[0] if criterios_usados else 'Popularidad'
    crit2 = criterios_usados[1] if len(criterios_usados) > 1 else criterios_usados[0] if criterios_usados else 'Ventas'
    
    # Obtener cortes de la nueva estructura
    A_cut_1 = cortes_abc.get(crit1, {}).get('A', 0.8) if cortes_abc else 0.8
    B_cut_1 = cortes_abc.get(crit1, {}).get('B', 0.95) if cortes_abc else 0.95
    A_cut_2 = cortes_abc.get(crit2, {}).get('A', 0.8) if cortes_abc else 0.8
    B_cut_2 = cortes_abc.get(crit2, {}).get('B', 0.95) if cortes_abc else 0.95

    # -------------------------------
    # Crear hoja Portada
    # -------------------------------
    # Crear datos de portada dinámicamente
    portada_campos = ['Documento leído', 'Hoja utilizada', 'Unidades de volumen', 'Criterios utilizados']
    portada_valores = [file_name, sheet_used, vol_units, ', '.join(criterios_usados)]
    
    # Agregar cortes para cada criterio
    for criterio in criterios_usados:
        if criterio in cortes_abc:
            portada_campos.extend([f'Corte A ({criterio})', f'Corte B ({criterio})'])
            portada_valores.extend([cortes_abc[criterio]['A'], cortes_abc[criterio]['B']])
    
    portada_data = {
        'Campo': portada_campos,
        'Valor': portada_valores
    }

    df_portada = pd.DataFrame(portada_data)
        
if 'by_item' in st.session_state:
    by_item = st.session_state['by_item']
    base = st.session_state['base']

    from statsmodels.tsa.seasonal import seasonal_decompose
    import plotly.graph_objects as go

    with st.expander("🔮 Análisis de Demanda por Artículo", expanded=False):
        st.header('🔮 Análisis de Demanda por Artículo')

        with st.expander("ℹ️ ¿Qué estudia esta sección?", expanded=False):
            st.markdown("""
            Esta sección está diseñada para **analizar la demanda histórica de un artículo específico**, permitiendo comprender cómo ha evolucionado a lo largo del tiempo y detectar patrones de comportamiento relevantes.  

    **Objetivos principales:**
    1. **Visualizar la serie histórica de demanda:**  
       - Observar la evolución de ventas o unidades despachadas.  
       - Analizar con frecuencia semanal o mensual según la granularidad de los datos.

    2. **Identificar la tendencia de la demanda:**  
       - Mediante suavizado (promedio móvil), se destaca la tendencia subyacente.  
       - Permite detectar si la demanda está **creciendo, decreciendo o se mantiene estable**.

    3. **Detectar estacionalidad o patrones repetitivos:**  
       - Analiza la autocorrelación de la serie para identificar ciclos de demanda.  
       - Los rezagos significativos indican posibles patrones estacionales.

    4. **Interpretación simplificada con pocos datos:**  
       - Para series cortas o irregulares, se ofrece un análisis alternativo con suavizado y autocorrelación parcial.
            """)

        with st.expander("ℹ️ ¿Qué métricas se utilizan y cómo interpretarlas?", expanded=False):
            st.markdown("""
             **Métricas y su interpretación:**
    
    1. **Serie histórica:**  
       - Gráfico de la demanda a lo largo del tiempo.  
       - Permite observar cambios, picos o caídas en las ventas.

    2. **Tendencia suavizada:**  
       - Promedio móvil de 3 períodos (o ajustable).  
       - Indica la dirección general de la demanda (creciente, decreciente o estable).

    3. **Autocorrelación (ACF):**  
       - Mide la correlación de la demanda con sus propios rezagos.  
       - **Interpretación de rezagos:**  
         - Un rezago de 1 mes/semana indica cómo la demanda actual se relaciona con la del período anterior.  
         - Rezagos significativos más largos (2, 3, …) muestran patrones repetitivos o estacionales.  
         - Por ejemplo, un rezago significativo de 12 meses sugiere que la demanda se repite anualmente en ese mes.

    4. **Interpretación resumida:**  
       - Con base en la tendencia y la ACF, se puede anticipar estacionalidad y dirección de la demanda.  
       - Útil para planificación de inventarios, producción y estrategias de abastecimiento.
            """)
        # Selección de nivel de análisis
        nivel_analisis = st.selectbox('Nivel de análisis', ['Articulo', 'Familia'], index=0)

        # Selección de artículo o familia
        if nivel_analisis == 'Articulo':
            opciones = sorted(base['Articulo'].unique())
            seleccion = st.selectbox('Selecciona Artículo para analizar', opciones, key='analizar_articulo')
            columna_estudio = 'Unidades' if st.selectbox('Unidad a pronosticar', ['Unidades vendidas', 'Cajas vendidas'], index=0) == 'Unidades vendidas' else 'Cajas_vendidas'
        else:
            opciones = sorted(base['Familia'].unique())
            seleccion = st.selectbox('Selecciona Familia para analizar', opciones, key='analizar_familia')
            columna_estudio = 'Unidades' if st.selectbox('Unidad a pronosticar', ['Unidades vendidas', 'Cajas vendidas'], index=0) == 'Unidades vendidas' else 'Cajas_vendidas'

        # Filtrar datos
        if nivel_analisis not in base.columns:
            st.error(f"El nivel de análisis '{nivel_analisis}' no existe en las columnas de base.")
            st.write("Columnas disponibles:", base.columns)
            st.stop()

        base_filtrada = base[base[nivel_analisis] == seleccion].copy()

        if base_filtrada.empty:
            st.warning(f"No hay registros para {nivel_analisis.lower()} seleccionado.")
            st.stop()

        # Serie histórica
        resample_freq = 'MS' if st.selectbox('Frecuencia de estudio', ['Mensual', 'Semanal'], index=0) == 'Mensual' else 'W-MON'
        ts = base_filtrada.groupby('Fecha')[columna_estudio].sum().resample(resample_freq).sum().fillna(0)
        ts.index.freq = resample_freq
        st.subheader("Serie histórica")
        st.line_chart(ts)

        # --- Panel interactivo para parámetros de análisis ---
        st.subheader("⚙️ Configuración del análisis")

        modelo_descomp = st.radio(
            "Modelo de descomposición",
            options=["Aditivo", "Multiplicativo"],
            index=0,
            horizontal=True
        )

        periodo_default = 12 if resample_freq == "MS" else 52
        periodo_estacionalidad = st.number_input(
            "Periodo de estacionalidad (ej. 12 meses, 52 semanas)",
            min_value=2, max_value=200, value=periodo_default, step=1
        )

        umbral_estacionalidad = st.slider(
            "Sensibilidad para considerar estacionalidad significativa",
            min_value=0.1, max_value=1.0, value=0.3, step=0.05
        )

        # --- Análisis de tendencia y estacionalidad ---
        st.subheader("📊 Tendencia y Estacionalidad")

        # Validar si hay suficientes datos
        if len(ts) >= 2 * periodo_estacionalidad:
            try:
                decomposition = seasonal_decompose(
                    ts,
                    model="additive" if modelo_descomp == "Aditivo" else "multiplicative",
                    period=periodo_estacionalidad
                )

                # Graficar resultados de la descomposición
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=decomposition.trend.index, y=decomposition.trend,
                                        mode='lines', name='Tendencia'))
                fig.add_trace(go.Scatter(x=decomposition.seasonal.index, y=decomposition.seasonal,
                                        mode='lines', name='Estacionalidad'))
                fig.add_trace(go.Scatter(x=decomposition.resid.index, y=decomposition.resid,
                                        mode='lines', name='Residuales'))

                fig.update_layout(
                    title="Descomposición de la serie",
                    xaxis_title="Fecha", yaxis_title="Demanda",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig, use_container_width=True)

                # Interpretación automática
                st.markdown("### 🔍 Interpretación automática")
                if decomposition.trend.notna().sum() > 0:
                    tendencia = decomposition.trend.dropna()
                    if tendencia.iloc[-1] > tendencia.iloc[0]:
                        st.success("🔼 La demanda muestra una **tendencia creciente**.")
                    elif tendencia.iloc[-1] < tendencia.iloc[0]:
                        st.error("🔽 La demanda muestra una **tendencia decreciente**.")
                    else:
                        st.info("⏸ La demanda se mantiene relativamente estable.")

                seasonal_strength = decomposition.seasonal.std() / ts.std() if ts.std() > 0 else 0
                if seasonal_strength > umbral_estacionalidad:
                    st.success("📈 Se observa una **estacionalidad significativa** (patrones recurrentes).")
                else:
                    st.info("📉 No se detecta una estacionalidad marcada en la serie.")

            except Exception as e:
                st.warning("⚠️ Ocurrió un problema en la descomposición.")
                st.error(e)

        else:
            # --- Análisis alternativo si no hay suficientes datos ---
            st.warning(f"""
            ⚠️ La serie solo tiene {len(ts)} observaciones, 
            pero se requieren al menos {2*periodo_estacionalidad} para descomponer.
            Se mostrará un análisis simplificado.
            """)

            # Rolling mean (suavizado de tendencia)
            rolling = ts.rolling(window=3, center=True).mean()
            st.line_chart(pd.DataFrame({
                "Demanda": ts, 
                "Tendencia suavizada (3 periodos)": rolling
            }))

            # Autocorrelación (ACF)
            from statsmodels.tsa.stattools import acf

            # Definir rango máximo de rezagos según frecuencia
            default_lag = 10
            if resample_freq == "MS":
                max_lag_default = min(12, len(ts)-1)  # hasta 12 meses
            else:  # Semanal
                max_lag_default = min(52, len(ts)-1)  # hasta 52 semanas

            # Slider interactivo para seleccionar número de rezagos
            max_lag = st.slider(
                "Número máximo de rezagos para ACF",
                min_value=5,
                max_value=max_lag_default,
                value=default_lag
            )

            lag_acf = acf(ts, nlags=max_lag)

            acf_df = pd.DataFrame({
                "Rezago": list(range(len(lag_acf))),
                "Autocorrelación": lag_acf
            })
            st.bar_chart(acf_df.set_index("Rezago"))

            # Interpretación simple de la tendencia
            st.markdown("### 🔍 Interpretación simplificada")
            if rolling.dropna().iloc[-1] > rolling.dropna().iloc[0]:
                st.success("🔼 La demanda parece **creciente** según el suavizado.")
            elif rolling.dropna().iloc[-1] < rolling.dropna().iloc[0]:
                st.error("🔽 La demanda parece **decreciente** según el suavizado.")
            else:
                st.info("⏸ La demanda parece **estable** en el tiempo analizado.")

            # Identificar rezagos significativos en ACF
            significativos = [lag for lag, val in enumerate(lag_acf[1:], start=1) if abs(val) > 0.3]

            if significativos:
                if resample_freq == "MS":
                    detalle = ", ".join([f"{lag} mes(es)" for lag in significativos])
                else:  # Semanal
                    detalle = ", ".join([f"{lag} semana(s)" for lag in significativos])

                st.success(f"📈 Se detectan correlaciones en los rezagos: **{detalle}**. "
                        "Esto sugiere una posible **estacionalidad** en esos intervalos.")
            else:
                st.info("📉 No se observa evidencia fuerte de estacionalidad en los rezagos analizados.")
                
    with st.expander("📊 Análisis por Familia", expanded=False):
        st.header("📊 Análisis por Familia")
        st.markdown("Analiza la demanda, rotaciones y ventas agrupadas por familia de productos.")

        # Agrupación por Familia
        familia_agg = base.groupby('Familia').agg(
            Ventas=('Monto', 'sum'),
            Unidades=('Unidades', 'sum'),
            Cajas=('Cajas_vendidas', 'sum'),
            Volumen=('Volumen_m3', 'sum'),
            Rotacion_Sem=('Cajas_vendidas', lambda x: x.sum() / weeks_range),
            Rotacion_Mes=('Cajas_vendidas', lambda x: x.sum() / months_range),
            Rotacion_Anual=('Cajas_vendidas', lambda x: x.sum() / years_range)
        ).reset_index()

        # Calcular porcentajes
        total_ventas = familia_agg['Ventas'].sum()
        familia_agg['% Ventas'] = (familia_agg['Ventas'] / total_ventas * 100).round(2)

        # Mostrar tabla
        st.dataframe(familia_agg)

        # Gráfico de contribución por familia
        fig_familia = px.pie(
            familia_agg,
            values='% Ventas',
            names='Familia',
            title="Contribución de Ventas por Familia"
        )
        st.plotly_chart(fig_familia, use_container_width=True)
    
    with st.expander("📊 Análisis por País", expanded=False):
        st.header("📊 Análisis por País")
        st.markdown("Analiza la demanda y ventas agrupadas por país (manteniendo anonimato).")

        # Crear columna TipoPais
        
        base['TipoPais'] = base['NumPais'].apply(lambda x: 'Nacional' if x == '01' else 'Exportación')

        # Agrupación por país
        pais_agg = base.groupby('NumPais').agg(
            Ventas=('Monto', 'sum'),
            Unidades=('Unidades', 'sum'),
            Cajas=('Cajas_vendidas', 'sum'),
            Volumen=('Volumen_m3', 'sum')
        ).reset_index()

        # Calcular porcentajes
        total_ventas = pais_agg['Ventas'].sum()
        pais_agg['% Ventas'] = (pais_agg['Ventas'] / total_ventas * 100).round(2)

        # Mostrar tabla
        st.dataframe(pais_agg)

        # Gráfico de contribución por país
        fig_pais = px.pie(
            pais_agg,
            names='NumPais',
            values='% Ventas',
            title="Contribución de Ventas por País",
        )
        st.plotly_chart(fig_pais, use_container_width=True)

               
    with st.expander("📊 Análisis de Contribución por Categorías ABC", expanded=False):
        st.header('📊 Análisis de Contribución por Categorías ABC')
        st.markdown("""
        Esta sección proporciona un análisis detallado de la contribución de cada categoría ABC 
        al total de ventas, volumen, popularidad, unidades y cajas, útil para entender el impacto de cada categoría.
        """)

        # --- Agrupar y calcular métricas ---
        contribucion_categorias = by_item.groupby('Clase_SuperABC').agg({
            'ventas': 'sum',
            'volumen': 'sum',
            'popularidad': 'sum',
            'Unidades': 'sum',
            'Cajas': 'sum',
            'Rot_Unidades_Sem': 'mean',
            'Rot_Unidades_Mes': 'mean',
            'Rot_Unidades_Año': 'mean',
            'Rot_Cajas_Sem': 'mean',
            'Rot_Cajas_Mes': 'mean',
            'Rot_Cajas_Año': 'mean'
        }).round(2)

        # Conteo de artículos
        contribucion_categorias['Cantidad_Articulos'] = by_item.groupby('Clase_SuperABC').size()

        # Totales para porcentajes
        total_ventas = contribucion_categorias['ventas'].sum()
        total_volumen = contribucion_categorias['volumen'].sum()
        total_popularidad = contribucion_categorias['popularidad'].sum()
        total_unidades = contribucion_categorias['Unidades'].sum()
        total_cajas = contribucion_categorias['Cajas'].sum()
        total_articulos = contribucion_categorias['Cantidad_Articulos'].sum()

        # Calcular porcentajes
        contribucion_categorias['% Ventas'] = (contribucion_categorias['ventas']/total_ventas*100).round(2)
        contribucion_categorias['% Volumen'] = (contribucion_categorias['volumen']/total_volumen*100).round(2)
        contribucion_categorias['% Popularidad'] = (contribucion_categorias['popularidad']/total_popularidad*100).round(2)
        contribucion_categorias['% Unidades'] = (contribucion_categorias['Unidades']/total_unidades*100).round(2)
        contribucion_categorias['% Cajas'] = (contribucion_categorias['Cajas']/total_cajas*100).round(2)
        contribucion_categorias['% Artículos'] = (contribucion_categorias['Cantidad_Articulos']/total_articulos*100).round(2)

        # Renombrar columnas
        contribucion_categorias.rename(columns={
            'ventas':'Ventas',
            'volumen':'Volumen',
            'popularidad':'Popularidad',
            'Unidades':'Unidades',
            'Cajas':'Cajas',
            'Rot_Unidades_Sem':'Rotación Semanal (uds)',
            'Rot_Unidades_Mes':'Rotación Mensual (uds)',
            'Rot_Unidades_Año':'Rotación Anual (uds)',
            'Rot_Cajas_Sem':'Rotación Semanal (cajas)',
            'Rot_Cajas_Mes':'Rotación Mensual (cajas)',
            'Rot_Cajas_Año':'Rotación Anual (cajas)'
        }, inplace=True)

        st.subheader("📈 Tabla de Contribución por Categorías")
        st.dataframe(contribucion_categorias)

        # --- Gráficos de contribución (ventas y volumen) ---
        import plotly.express as px
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("🥧 Contribución de Ventas por Categoría")
            fig_ventas = px.pie(
                contribucion_categorias,
                values='% Ventas',
                names=contribucion_categorias.index,
                title="Distribución de Ventas por Categoría ABC"
            )
            st.plotly_chart(fig_ventas, use_container_width=True)

        with col2:
            st.subheader("🥧 Contribución de Volumen por Categoría")
            fig_volumen = px.pie(
                contribucion_categorias,
                values='% Volumen',
                names=contribucion_categorias.index,
                title="Distribución de Volumen por Categoría ABC"
            )
            st.plotly_chart(fig_volumen, use_container_width=True)

        # --- Gráficos de rotación ---
        st.subheader("🥧 Distribución de Rotación por Categoría")
        periodos = ['Semanal','Mensual','Anual']
        tipos = ['uds','cajas']

        for tipo in tipos:
            for periodo in periodos:
                col_name = f'Rotación {periodo} ({tipo})'
                if col_name in contribucion_categorias.columns:
                    fig = px.pie(
                        contribucion_categorias,
                        values=col_name,
                        names=contribucion_categorias.index,
                        title=f"Distribución de Rotación {periodo} ({tipo})"
                    )
                    st.plotly_chart(fig, use_container_width=True)

        # --- Gráfico de barras comparativo ---
        st.subheader("📊 Comparación de Contribuciones por Métricas")
        contrib_data_melted = contribucion_categorias[[
            '% Ventas','% Volumen','% Popularidad','% Unidades','% Cajas'
        ]].reset_index().melt(id_vars=['Clase_SuperABC'], 
                            var_name='Métrica', value_name='Porcentaje')

        fig_barras = px.bar(
            contrib_data_melted,
            x='Clase_SuperABC',
            y='Porcentaje',
            color='Métrica',
            barmode='group',
            title="Comparación de Contribuciones por Categoría ABC"
        )
        fig_barras.update_layout(xaxis_title="Categoría ABC", yaxis_title="Porcentaje (%)")
        st.plotly_chart(fig_barras, use_container_width=True)

        # --- Insights automáticos ---
        st.subheader("💡 Insights Automáticos")
        insights = []

        # Ventas y concentración
        top_categoria = contribucion_categorias['% Ventas'].idxmax()
        top_ventas = contribucion_categorias.loc[top_categoria,'% Ventas']
        insights.append(f"🎯 **Categoría líder:** {top_categoria} representa {top_ventas}% de las ventas totales")

        categorias_80 = contribucion_categorias[contribucion_categorias['% Ventas'].cumsum() <= 80]
        insights.append(f"📊 **Concentración 80/20:** {len(categorias_80)} categorías concentran 80% de las ventas")

        categorias_bajas = contribucion_categorias[contribucion_categorias['% Ventas'] < 5]
        if len(categorias_bajas) > 0:
            insights.append(f"📉 **Baja contribución:** {len(categorias_bajas)} categorías <5% de ventas")
            insights.append(f"🔍 **Revisar categorías:** {', '.join(categorias_bajas.index)}")

        # Top rotaciones
        for tipo in tipos:
            for periodo in periodos:
                col_name = f'Rotación {periodo} ({tipo})'
                if col_name in contribucion_categorias.columns:
                    top_rot = contribucion_categorias.sort_values(col_name, ascending=False).head(3)
                    insights.append(f"⚡ **Mayor rotación {periodo} ({tipo}):** {', '.join(top_rot.index)}")

        for insight in insights:
            st.info(insight)


    with st.expander("🏭 Sugerencias de Distribución de Bodega", expanded=False):
        # -------------------------------
        # Sugerencias de Distribución de Bodega
        # -------------------------------
        st.header('🏭 Sugerencias de Distribución de Bodega')
        
        st.markdown("""
        Esta sección permite calcular la distribución óptima de racks en la bodega basándose en el análisis ABC 
        y las dimensiones físicas de pallets, bays, racks y la bodega. El sistema calcula automáticamente 
        la capacidad de almacenamiento y sugiere la distribución de racks por categoría ABC.
        
        **💡 Nota importante:** El porcentaje de almacenamiento representa qué parte del área total de la bodega 
        se destinará específicamente para almacenamiento (el resto se usa para pasillos, oficinas, áreas de 
        recepción/despacho, etc.). Un valor típico es entre 60-80%.
        """)
        
        # Verificar que tenemos datos de Súper ABC
        if 'Clase_SuperABC' not in by_item.columns:
            st.warning("⚠️ Primero debes calcular el Súper ABC para usar esta funcionalidad.")
        else:
            # Crear columnas para organizar los parámetros
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📦 Dimensiones de Pallets (metros)")
                largo_pallet = st.number_input("Largo del pallet (m)", min_value=0.1, value=1.2, step=0.01)
                ancho_pallet = st.number_input("Ancho del pallet (m)", min_value=0.1, value=1.0, step=0.01)
                alto_pallet = st.number_input("Alto del pallet (m)", min_value=0.1, value=1.5, step=0.01)
                factor_llenado = st.number_input("Factor de llenado (%)", min_value=0.1, max_value=100.0, value=85.0, step=1.0) / 100.0

                st.subheader("🏗️ Dimensiones de Bay (metros)")
                largo_bay = st.number_input("Largo de Bay (m)", min_value=0.1, value=2.5, step=0.01)
                profundidad_bay = st.number_input("Profundidad del Bay (m)", min_value=0.1, value=1.2, step=0.01)
                niveles = st.number_input("Niveles", min_value=1, value=5, step=1)
                
            with col2:
                st.subheader("🏢 Información de Rack")
                bays_por_rack = st.number_input("Bays por rack", min_value=1, value=10, step=1)

                st.subheader("🏭 Dimensiones de Bodega (metros)")
                ancho_bodega = st.number_input("Ancho de bodega (m)", min_value=0.1, value=30.0, step=0.01)
                largo_bodega = st.number_input("Largo de bodega (m)", min_value=0.1, value=60.0, step=0.01)
                porcentaje_almacenamiento = st.number_input("Porcentaje para almacenamiento (%)", min_value=1.0, max_value=100.0, value=70.0, step=1.0) / 100.0
                ancho_pasillo = st.number_input("Ancho de pasillo (m)", min_value=0.1, value=3.6, step=0.01)
            
            # Botón para calcular distribución
            if st.button("🧮 Calcular Distribución de Bodega"):
                
                # -------------------------------
                # Cálculos de Capacidad en metros
                # -------------------------------
                st.subheader("📊 Cálculos de Capacidad")

                # Volumen de pallet (m³)
                volumen_pallet = largo_pallet * ancho_pallet * alto_pallet * factor_llenado

                # Pallets por nivel
                pallets_por_nivel = int((largo_bay // largo_pallet) * (profundidad_bay // ancho_pallet))

                # Área de bay (m²)
                area_bay = largo_bay * profundidad_bay

                # Área de rack (m²)
                area_rack = area_bay * bays_por_rack

                # Área de bodega total (m²)
                area_bodega_total = ancho_bodega * largo_bodega

                # Área efectiva de almacenamiento (m²)
                area_efectiva_almacenamiento = area_bodega_total * porcentaje_almacenamiento

                # Área de pasillo (m²)
                area_pasillo = largo_bay * ancho_pasillo * bays_por_rack

                # Área de rack + pasillo (m²)
                area_rack_pasillo = area_rack + area_pasillo

                # Pallets por rack
                pallets_por_rack = pallets_por_nivel * niveles * bays_por_rack
                
                
                # Mostrar cálculos
                calculos_df = pd.DataFrame({
                    'Métrica': [
                        'Volumen de pallet (m³)',
                        'Pallets por nivel',
                        'Área de bay (m²)',
                        'Área de rack (m²)',
                        'Área de bodega (m²)',
                        f'Área efectiva de almacenamiento (m²) - {porcentaje_almacenamiento*100:.0f}%',
                        'Área de pasillo (m²)',
                        'Área de rack + pasillo (m²)',
                        'Pallets por rack'
                    ],
                    'Valor': [
                        round(volumen_pallet, 3),
                        pallets_por_nivel,
                        round(area_bay, 2),
                        round(area_rack, 2),
                        round(area_bodega_total, 2),
                        round(area_efectiva_almacenamiento, 2),
                        round(area_pasillo, 2),
                        round(area_rack_pasillo, 2),
                        pallets_por_rack
                    ]
                })
                st.dataframe(calculos_df, use_container_width=True)

                
                # -------------------------------
                # Tabla previa de preparación (solicitada)
                # -------------------------------
                st.subheader("📋 Tabla base previa (detalle limpio)")
                base_pre = base.copy()
                base_pre['MesyAño'] = base_pre['Fecha'].dt.to_period('M').astype(str)
                # Volumen por unidad: cuidar divisiones por 0
                base_pre['Volumen por unidad'] = (base_pre['Volumen_m3'] / base_pre['Unidades'].replace(0, np.nan)).fillna(0)
                tabla_previa = base_pre[['NumDoc','Articulo','Unidades','Cajas_vendidas','Unidades_por_caja','Fecha','Volumen_m3','MesyAño','Volumen por unidad']].copy()
                tabla_previa.columns = ['Num. Doc','Artículo','Unid. Vend','Cajas vend.','Cant x caja','Fecha Doc','Volumen total (p3)','MesyAño','Volumen por unidad']
                st.dataframe(tabla_previa, use_container_width=True)

                # -------------------------------
                # Análisis de Inventario por SKU (Artículo)
                # -------------------------------
                st.subheader("📋 Análisis por SKU (Artículo)")

                # Utilidades
                SCORES = {
                "A": 0.99,
                "B": 0.90,
                "C": 0.75
                }

                def calcular_csl(categoria):
                    if not isinstance(categoria, str):
                        return 0.70
                    valores = [SCORES.get(letra, 0.70) for letra in categoria] # Define un score por cada letra, y el CSL de la combinación es el promedio
                    return round(sum(valores) / len(valores), 2)


                def calcular_z_score(csl):
                    from scipy.stats import norm
                    return float(norm.ppf(csl))
                
                def calcular_safety_stock(demanda_promedio, desviacion, z_score, lead_time=1):
                    return z_score * desviacion * np.sqrt(lead_time)
                
                # 1) Tabla mensual de Unidades por SKU
                base_mes = base.copy()
                base_mes['YearMonth'] = base_mes['Fecha'].dt.to_period('M').astype(str)
                pivot_mes = pd.pivot_table(
                    base_mes,
                    index='Articulo',
                    columns='YearMonth',
                    values='Unidades',
                    aggfunc='sum',
                    fill_value=0
                )
                # ordenar columnas cronológicamente
                pivot_mes = pivot_mes.reindex(sorted(pivot_mes.columns), axis=1)

                # 2) Cálculos por SKU: totales y estadísticas de demanda
                totales = base.groupby('Articulo').agg(
                    unidades_totales=('Unidades','sum'),
                    volumen_total=('Volumen_m3','sum')
                )
                vol_por_unidad = (totales['volumen_total'] / totales['unidades_totales'].replace(0, np.nan)).fillna(0)

                # Unidades por caja por SKU: usar la moda (valor más frecuente) de 'Cant x Caja'
                upc_series = pd.to_numeric(base['Unidades_por_caja'], errors='coerce')
                upc_por_sku = base.assign(Unidades_por_caja=upc_series).groupby('Articulo')['Unidades_por_caja'].agg(
                    lambda s: s.mode().iloc[0] if not s.mode().empty else s.dropna().iloc[0] if not s.dropna().empty else 1
                )
                upc_por_sku = upc_por_sku.fillna(1).replace(0, 1)

                # Mapear categoría Súper ABC por SKU EXACTAMENTE como fue calculada (sin reordenar letras)
                # Usar by_item del session_state que contiene las categorías ABC originales
                by_item_original = st.session_state['by_item']
                mapa_abc = by_item_original['Clase_SuperABC'].reindex(pivot_mes.index)
                # Filtrar solo SKUs que tienen clasificación ABC válida
                skus_con_abc = mapa_abc.dropna()
                pivot_mes_filtrado = pivot_mes.reindex(skus_con_abc.index)
                
                # Demanda promedio y desviación: usar promedio mensual directo del pivote filtrado
                demanda_prom = pivot_mes_filtrado.mean(axis=1)
                # Desviación muestral (como DESVEST.M en Excel)
                desviacion = pivot_mes_filtrado.std(axis=1, ddof=1).fillna(0)
                
                csl = skus_con_abc.apply(calcular_csl)
                z_vals = csl.apply(calcular_z_score)
                ss_vals = calcular_safety_stock(demanda_prom, desviacion, z_vals)
                inv_max = demanda_prom + ss_vals

                # Volúmenes y cajas (usar solo SKUs con ABC válido)
                vol_por_caja = vol_por_unidad.reindex(skus_con_abc.index).fillna(0) * upc_por_sku.reindex(skus_con_abc.index).fillna(1)
                vol_total_unidades = inv_max * vol_por_unidad.reindex(skus_con_abc.index).fillna(0)
                cant_cajas = np.ceil(inv_max / upc_por_sku.reindex(skus_con_abc.index).replace(0, 1))
                vol_total_cajas = cant_cajas * vol_por_caja

                # Construir tabla final por SKU (solo con SKUs que tienen ABC)
                sku_df = pivot_mes_filtrado.copy()
                sku_df['Total general'] = pivot_mes_filtrado.sum(axis=1)
                sku_df['Demanda promedio'] = demanda_prom.round(2)
                sku_df['Desviación'] = desviacion.round(2)
                sku_df['ABC'] = skus_con_abc
                sku_df['CSL'] = csl
                sku_df['Z'] = z_vals.round(2)
                sku_df['ss'] = ss_vals.round(2)
                sku_df['Inventario máximo'] = inv_max.round(2)
                sku_df['Volumen por unidad (m³)'] = vol_por_unidad.reindex(skus_con_abc.index).fillna(0).round(4)
                sku_df['Volumen Total (unidades, m³)'] = vol_total_unidades.round(2)
                sku_df['Unidades por caja'] = upc_por_sku.reindex(skus_con_abc.index).fillna(1).astype(int)
                sku_df['Cantidad de cajas'] = cant_cajas.astype(int)
                sku_df['Volumen por caja (m³)'] = vol_por_caja.round(4)
                sku_df['Volumen Total (cajas, m³)'] = vol_total_cajas.round(2)

                st.dataframe(sku_df.reset_index(), use_container_width=True)

                # Totales generales previos a demanda: sumar meses, Total general y Volumen Total (cajas)
                totales_previos = {
                    'Total general (unidades)': float(sku_df['Total general'].sum()),
                    'Volumen Total (cajas, m³)': float(sku_df['Volumen Total (cajas, m³)'].sum())
                }
                st.write('Totales generales:', {k: round(v, 2) for k, v in totales_previos.items()})
                
                # -------------------------------
                # Tabla de Volúmenes por Categoría (antes de racks)
                # -------------------------------
                st.subheader("📊 Volúmenes por Categoría ABC")
                
                # Agregar volúmenes por categoría (sku_df ya está filtrado por ABC válido)
                vol_por_categoria = sku_df.groupby('ABC')['Volumen Total (cajas, m³)'].sum().sort_index()
                vol_total_general = vol_por_categoria.sum()
                
                tabla_volumenes = pd.DataFrame({
                    'Suma de Volumen Total': vol_por_categoria,
                    'Porcentaje del Total': (vol_por_categoria / vol_total_general * 100)
                }).round(4)

                # % agrupado por primera letra
                tabla_volumenes['Primera_Letra'] = tabla_volumenes.index.astype(str).str[0]
                pct_grouped = (tabla_volumenes.groupby('Primera_Letra')['Porcentaje del Total'].sum()).round(4)
                tabla_volumenes['Porcentaje agrupado'] = tabla_volumenes['Primera_Letra'].map(pct_grouped)

                # Añadir Total General
                total_row = pd.DataFrame({
                    'Suma de Volumen Total': [vol_total_general],
                    'Porcentaje del Total': [100.0],
                    'Porcentaje agrupado': [pct_grouped.sum()]
                }, index=['Total General'])
                tabla_mostrar = pd.concat([tabla_volumenes[['Suma de Volumen Total', 'Porcentaje del Total', 'Porcentaje agrupado']].round(2), total_row], axis=0)
                tabla_mostrar.index.name = 'Etiquetas de fila'
                st.dataframe(tabla_mostrar, use_container_width=True)

                st.dataframe(vol_por_categoria.to_frame('Volumen Total (cajas)').round(2), use_container_width=True)
                
                # -------------------------------
                # Cálculo de Racks Necesarios
                # -------------------------------
                st.subheader("🏗️ Cálculo de Racks Necesarios")
                
                # Crear tabla de racks por categoría usando los volúmenes ya calculados
                racks_categorias = pd.DataFrame({'VolumenTotal_m3': vol_por_categoria})
                racks_categorias.index.name = 'Clase_SuperABC'
                
                # Cálculos paso a paso 
                racks_categorias['Pallets_Necesarios'] = np.ceil(racks_categorias['VolumenTotal_m3'] / volumen_pallet)
                racks_categorias['Equivalente_en_Niveles'] = racks_categorias['Pallets_Necesarios'] / pallets_por_nivel
                racks_categorias['Equivalente_en_Bays'] = racks_categorias['Equivalente_en_Niveles'] / niveles
                racks_categorias['Equivalente_en_Racks'] = racks_categorias['Equivalente_en_Bays'] / bays_por_rack
                
                # Calcular distribución de racks en porcentaje respecto a racks redondeados (espacio no usado)
                total_racks = racks_categorias['Equivalente_en_Racks'].sum()
                if total_racks == 0:
                    racks_categorias['Distribucion_de_Racks_%'] = 0.0
                else:
                    racks_disponibles = np.ceil(total_racks)
                    racks_categorias['Distribucion_de_Racks_%'] = (racks_categorias['Equivalente_en_Racks'] / racks_disponibles * 100).round(2)

                    # Calcular sobrante
                    sobrante_pct = 100 - racks_categorias['Distribucion_de_Racks_%'].sum()
                    if sobrante_pct > 0:
                        sobrante_row = pd.DataFrame({
                            'VolumenTotal_m3': [0],
                            'Pallets_Necesarios': [0],
                            'Equivalente_en_Niveles': [0],
                            'Equivalente_en_Bays': [0],
                            'Equivalente_en_Racks': [racks_disponibles - total_racks],
                            'Distribucion_de_Racks_%': [sobrante_pct]
                        }, index=['Sobrante'])
                        racks_categorias = pd.concat([racks_categorias, sobrante_row])

                # Mostrar tabla de racks
                st.dataframe(racks_categorias, use_container_width=True)
                
                # -------------------------------
                # Resumen de Distribución
                # -------------------------------
                st.subheader("📊 Resumen de Distribución")
                
                # Calcular métricas de utilización
                racks_necesarios = np.ceil(racks_categorias['Pallets_Necesarios'].sum() / pallets_por_rack)
                area_en_uso = area_rack_pasillo * racks_necesarios
                utilizacion_espacio = (area_en_uso / area_efectiva_almacenamiento * 100).round(2)
                
                resumen_df = pd.DataFrame({
                    'Métrica': [
                        'Racks necesarios total',
                        'Área en uso (m²)',
                        'Utilización de espacio (%)'
                    ],
                    'Valor': [
                        int(racks_necesarios),
                        round(area_en_uso, 2),
                        utilizacion_espacio
                    ]
                })
                
                st.dataframe(resumen_df, use_container_width=True)
                
                # -------------------------------
                # Gráfico de Distribución
                # -------------------------------
                st.subheader("📊 Gráfico de Distribución de Racks por Categoría")
                
                fig_distribucion = px.pie(
                    values=racks_categorias['Distribucion_de_Racks_%'],
                    names=racks_categorias.index,
                    title="Distribución de Racks por Categoría ABC"
                )
                st.plotly_chart(fig_distribucion, use_container_width=True)
                
                # -------------------------------
                # Guardar resultados en session_state
                # -------------------------------
                st.session_state['calculos_capacidad'] = calculos_df
                st.session_state['analisis_categorias'] = sku_df
                st.session_state['racks_categorias'] = racks_categorias
                st.session_state['resumen_distribucion'] = resumen_df               
                st.success("✅ Distribución de bodega calculada exitosamente!")


# =============================================================================
# SECCIÓN DE DESCARGAS
# =============================================================================

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from openpyxl.chart.axis import DateAxis, NumericAxis

st.markdown("---")
st.header("📥 Descargas y Exportación")

with st.expander("📊 Descargar Resultados Completos", expanded=False):
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📈 Perfiles de Actividad")
        if st.session_state.get('want_csv', True):
            if st.button("📥 Exportar perfiles a Excel"):
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    # Hoja Portada primero
                    df_portada.to_excel(writer, sheet_name='Portada', index=False)

                    # Guardamos nombres de hojas para generar gráficos luego
                    hoja_nombres = []

                    for key, df in st.session_state.items():
                        if key.startswith("perfil_") or key == "perfil_by_item":
                            hoja = key.replace("perfil_", "")[:30]  # hoja ≤ 31 chars
                            # ✅ Si el DataFrame tiene índice con nombre, lo pasamos a columna
                            if df.index.name is not None:
                                df = df.reset_index()
                            df.to_excel(writer, sheet_name=hoja, index=False)
                            hoja_nombres.append((hoja, df))

                # Abrir libro para añadir gráficos
                buffer.seek(0)
                wb = load_workbook(buffer)

                for hoja, df in hoja_nombres:
                    ws = wb[hoja]
                    chart = BarChart()  # default, lo cambiaremos según hoja

                    # Selección de datos según tipo de hoja
                    if hoja.lower() in ["dias", "lineas", "carga", "cubicaje"]:
                        # Columna 1 = etiquetas, columna 2 = valores
                        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        chart.title = f"Gráfico {hoja}"
                        chart.y_axis.title = "Cantidad"
                        chart.x_axis.title = "Categoría"
                        ws.add_chart(chart, "H5")  # colocar gráfico a la derecha
                    elif hoja.lower() == "pareto":
                        chart = LineChart()
                        chart.title = "Pareto de Picks"
                        chart.style = 10  # estilo opcional
                        chart.legend = None  # sin leyenda
                        
                        # Datos Y = cum_pct_picks (columna C)
                        data = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
                        chart.add_data(data, titles_from_data=False)
                        
                        # Eje X = pct_sku (columna E)
                        cats = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
                        chart.set_categories(cats)
                        
                        # Etiquetas de ejes
                        chart.x_axis.title = "% acumulado de SKU"
                        chart.y_axis.title = "% acumulado de Picks"
                        
                        # Configurar líneas de graduación cada 10%
                        chart.x_axis.majorUnit = 10
                        chart.x_axis.minorUnit = 5
                        chart.y_axis.majorUnit = 10
                        chart.y_axis.minorUnit = 5
                        
                        # Posición del gráfico en la hoja
                        ws.add_chart(chart, "H5")
                    else:
                        continue  # saltar hojas no reconocidas
                    
                # Guardar cambios
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    "📊 Descargar Excel con perfiles",
                    data=buffer.getvalue(),
                    file_name="perfiles_distribuciones.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("No hay perfiles generados para descargar. Marca la opción 'Generar Excel' y vuelve a calcular.")
    with col2:
        st.subheader("🏗️ Distribución de Bodega")
        if st.session_state.get('want_csv', True):
            if st.button("📥 Exportar análisis de bodega", key="download_warehouse"):
                if 'analisis_categorias' in st.session_state:
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                        # Hoja Análisis por SKU
                        st.session_state['analisis_categorias'].to_excel(writer, sheet_name='Analisis_SKU', index=True)
                        
                        # Hoja Cálculos de Capacidad
                        if 'calculos_capacidad' in st.session_state:
                            st.session_state['calculos_capacidad'].to_excel(writer, sheet_name='Calculos_Capacidad', index=True)
                    
                        # Hoja Racks por Categoría
                        if 'racks_categorias' in st.session_state:
                            st.session_state['racks_categorias'].to_excel(writer, sheet_name='Distribucion_Racks', index=True)
                        
                        # Hoja Resumen de Distribución
                        if 'resumen_distribucion' in st.session_state:
                            st.session_state['resumen_distribucion'].to_excel(writer, sheet_name='Resumen_Distribucion', index=False)


                    from openpyxl import load_workbook
                    # 🔹 Cargar el libro desde el mismo buffer
                    buffer.seek(0)
                    wb = load_workbook(buffer)

                    from openpyxl.chart import PieChart, Reference
                    from openpyxl.chart.label import DataLabelList


                    # 🔹 Insertar gráfico en hoja Distribucion_Racks
                    if 'Distribucion_Racks' in wb.sheetnames:
                        ws2 = wb['Distribucion_Racks']
                        chart2 = PieChart()
                        labels2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)  # categorías
                        data2 = Reference(ws2, min_col=7, min_row=1, max_row=ws2.max_row)    # columna con % racks
                        chart2.add_data(data2, titles_from_data=True)
                        chart2.set_categories(labels2)
                        chart2.title = "Distribución de Racks por Categoría"
                        chart2.dataLabels = DataLabelList()
                        chart2.dataLabels.showVal = True      # Muestra valores
                        chart2.dataLabels.showPercent = True  # (opcional) muestra % en vez de valores
                        chart2.dataLabels.showCatName = True  # (opcional) muestra nombre de categoría
                        ws2.add_chart(chart2, "H5")  # posición del gráfico

                    # 🔹 Guardar de nuevo en buffer
                    new_buffer = io.BytesIO()
                    wb.save(new_buffer)
                    new_buffer.seek(0)

                    st.download_button(
                        "📥 Descargar Excel de Bodega",
                        data=new_buffer.getvalue(),
                        file_name='distribucion_bodega.xlsx',
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                else:
                    st.warning("Primero debes calcular la distribución de bodega")

with st.expander("📄 Reportes PDF", expanded=False):
    # -------------------------------
    # Generar PDF completo robusto y profesional (mejorado)
    # -------------------------------
    if st.session_state.get('gen_pdf', False):
        if not PDF_LIBS_AVAILABLE:
            st.error('Para generar PDFs instala: pip install reportlab matplotlib')
        else:
            if st.button('4) Generar informe PDF'):
                from reportlab.lib import colors
                from reportlab.platypus import TableStyle, Image, PageBreak
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import cm, mm
                from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
                import io, matplotlib.pyplot as plt

                # --- Pie de página con numeración
                def add_page_number(canvas, doc):
                    page_num = canvas.getPageNumber()
                    text = f"Página {page_num}"
                    canvas.setFont('Helvetica', 8)
                    canvas.drawRightString(200*mm, 10*mm, text)

                buffer = io.BytesIO()
                doc = BaseDocTemplate(
                    buffer,
                    pagesize=letter,
                    rightMargin=25, leftMargin=25,
                    topMargin=25, bottomMargin=18
                )
                frame = Frame(doc.leftMargin, doc.bottomMargin,
                            doc.width, doc.height, id='normal')
                template = PageTemplate(id='with-number',
                                        frames=frame,
                                        onPage=add_page_number)
                doc.addPageTemplates([template])

                styles = getSampleStyleSheet()
                elems = []

                # -------------------------------
                # Encabezado
                # -------------------------------
                elems.append(Paragraph('📊 Informe de Análisis - Súper ABC & Perfiles', styles['Title']))
                elems.append(Spacer(1, 14))

                # -------------------------------
                # Texto explicativo inicial
                # -------------------------------
                intro_text = """
                <b>Clasificación de zonas de bodega:</b><br/>
                - <b>Zona Oro (Close to door, close to floor):</b> Área de mayor valor, ubicada estratégicamente cerca de las puertas de entrada y salida de la bodega. Se destina a los productos de <b>alta rotación</b>, minimizando tiempo de viaje y esfuerzo de los operarios.<br/>
                - <b>Zona Plata (Close to floor):</b> Ubicada a una distancia media de las puertas. Se utiliza para productos de <b>rotación media</b>. El tiempo de acceso es moderado.<br/>
                - <b>Zona Bronce (Far from door, far from floor):</b> Área más alejada de las puertas. Reservada para productos de <b>baja rotación</b>. Aunque implica mayor tiempo de acceso, la baja frecuencia de movimiento lo justifica.<br/><br/>

                <b>Políticas de inventario:</b><br/>
                - <b>ROP-OUL:</b> Reordenar al alcanzar el punto de pedido (ROP), con un límite superior (OUL) para evitar exceso de inventario.<br/>
                - <b>RTP-EOQ:</b> Política de revisión periódica (RTP), aplicando el tamaño de lote económico (EOQ) como cantidad óptima de pedido.<br/>
                - <b>ROP-EOQ:</b> Política de reorden continuo (ROP), usando el EOQ como lote de reposición.<br/><br/>

                <b>Fill rate:</b> Métrica de nivel de servicio que mide el porcentaje de demanda atendida en el primer intento con el inventario disponible. Un fill rate alto indica capacidad de satisfacer pedidos sin generar faltantes.<br/><br/>

                <b>IRA (Inventory Record Accuracy):</b> KPI que mide la exactitud del inventario, comparando los registros teóricos del sistema con la realidad física del stock disponible en un almacén. Un IRA alto indica que la información del sistema es confiable, lo que permite una gestión de inventarios más eficiente, reduciendo pérdidas, excedentes y retrasos en los pedidos.  <br/><br/>

                <b>Recuento cíclico:</b> Estrategia de control de inventarios que consiste en revisar y contar de forma periódica subgrupos de productos a lo largo del año. Se enfoca más en artículos críticos o de mayor rotación (categoría A o AA), garantizando precisión de inventario sin necesidad de inventarios generales completos.
                """

                elems.append(Paragraph(intro_text, styles['Normal']))
                elems.append(Spacer(1, 14))

                # -------------------------------
                # Datos generales
                # -------------------------------
                file_name = st.session_state.get('file_name', uploaded_file.name if uploaded_file else 'Archivo no registrado')
                sheet_used = st.session_state.get('sheet_name', sheet_name or 'Hoja no registrada')
                vol_units = st.session_state.get('vol_units', unit_vol)
                criterios_usados = st.session_state.get('criterios_seleccionados', [crit1, crit2])
                cortes_abc = st.session_state.get('cortes_abc', {})
                
                # Para compatibilidad con código existente
                crit1 = criterios_usados[0] if criterios_usados else 'Popularidad'
                crit2 = criterios_usados[1] if len(criterios_usados) > 1 else criterios_usados[0] if criterios_usados else 'Ventas'
                
                # Obtener cortes de la nueva estructura
                A_cut_1 = cortes_abc.get(crit1, {}).get('A', 0.8) if cortes_abc else 0.8
                B_cut_1 = cortes_abc.get(crit1, {}).get('B', 0.95) if cortes_abc else 0.95
                A_cut_2 = cortes_abc.get(crit2, {}).get('A', 0.8) if cortes_abc else 0.8
                B_cut_2 = cortes_abc.get(crit2, {}).get('B', 0.95) if cortes_abc else 0.95

                general_info = f"""
                <b>Documento leído:</b> {file_name}<br/>
                <b>Hoja utilizada:</b> {sheet_used}<br/>
                <b>Unidades de volumen:</b> {vol_units}<br/>
                <b>Criterios utilizados:</b> {', '.join(criterios_usados)}<br/>
                """
                
                # Agregar cortes para cada criterio
                for criterio in criterios_usados:
                    if criterio in cortes_abc:
                        general_info += f"<b>Corte A ({criterio}):</b> {cortes_abc[criterio]['A']*100:.1f}%<br/>"
                        general_info += f"<b>Corte B ({criterio}):</b> {cortes_abc[criterio]['B']*100:.1f}%<br/>"
                elems.append(Paragraph(general_info, styles['Normal']))
                elems.append(Spacer(1, 12))

                by_item = st.session_state['by_item']
                base = st.session_state['base']

                # -------------------------------
                # Tabla resumen Super ABC (columnas compactas)
                # -------------------------------
                summary_table = by_item.groupby('Clase_SuperABC').agg(
                    Cantidad=('Clase_SuperABC','count'),
                    Zona_Bodega=('Zona_Bodega','first'),
                    Politica=('Política_Inv','first'),
                    FillRate=('FillRate_obj','first'),
                    Frecuencia_Recuento=('Frecuencia_Recuento','first'),
                    Ventas=('ventas','sum')
                ).reset_index()

                summary_table['Porcentaje'] = (summary_table['Cantidad']/summary_table['Cantidad'].sum()*100).round(2)
                total_sales = summary_table['Ventas'].sum()
                summary_table['% Ventas'] = (100 * summary_table['Ventas'] / (total_sales if total_sales>0 else 1)).round(2)
                summary_table['Ventas'] = summary_table['Ventas'].round(2)

                # 👉 Definir IRA según categoría usando la nueva función
                summary_table['IRA'] = summary_table['Clase_SuperABC'].apply(ira_by_class)

                # Reordenar columnas para poner IRA después de FillRate
                cols = list(summary_table.columns)
                insert_pos = cols.index('FillRate') + 1
                cols = cols[:insert_pos] + ['IRA'] + cols[insert_pos:-1]  # dejamos % Ventas al final
                summary_table = summary_table[cols]

                # preparar datos y anchos
                data = [list(summary_table.columns)] + summary_table.round(2).astype(str).values.tolist()
                col_widths = []
                for col in summary_table.columns:
                    if col in ['Cantidad','Zona_Bodega','FillRate','IRA']:
                        col_widths.append(45)
                    elif col in ['Ventas','Porcentaje','% Ventas']:
                        col_widths.append(50)
                    elif col in ['Clase_SuperABC','Frecuencia_Recuento']:
                        col_widths.append(70)
                    else:
                        col_widths.append(97)

                t = Table(data, colWidths=col_widths, hAlign='CENTER')
                t.setStyle(TableStyle([
                    ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('ALIGN',(0,0),(-1,-1),'CENTER')
                ]))
                elems.append(Paragraph('📑 Resumen por categoría (AA..CC)', styles['Heading2']))
                elems.append(t)
                elems.append(PageBreak())

                # -------------------------------
                # Función auxiliar para añadir figuras
                # -------------------------------
                def add_fig(fig, title='', width=450, height=240):
                    img_buf = io.BytesIO()
                    fig.tight_layout()
                    fig.savefig(img_buf, format='png', dpi=130)
                    plt.close(fig)
                    img_buf.seek(0)
                    elems.append(Paragraph(title, styles['Heading3']))
                    elems.append(Image(img_buf, width=width, height=height))
                    elems.append(Spacer(1, 12))
                # -------------------------------
                # Gráfica Pareto
                # -------------------------------

                pareto = by_item.sort_values('popularidad', ascending=False).copy()
                pareto['cum_picks'] = pareto['popularidad'].cumsum()
                total_picks = pareto['popularidad'].sum()
                pareto['cum_pct_picks'] = 100*pareto['cum_picks']/(total_picks if total_picks>0 else 1)
                pareto['pct_sku'] = 100 * np.arange(1,len(pareto)+1)/len(pareto)
                fig1, ax1 = plt.subplots(figsize=(6,3))
                ax1.plot(pareto['pct_sku'], pareto['cum_pct_picks'], marker='o')
                ax1.set_xlabel('% SKU (acumulado)')
                ax1.set_ylabel('% picks (acumulado)')
                ax1.set_title('Distribución de popularidad')
                add_fig(fig1, 'Pareto de popularidad')
                            
                pareto_intro = """
                Este perfil muestra qué porcentaje acumulado de los movimientos de picking corresponde a qué porcentaje acumulado de SKUs según el principio de Pareto (muchos triviales, pocos vitales). 
                Permite identificar los productos que concentran la mayor parte de la actividad y que deben recibir prioridad en la bodega.
                """
                elems.append(Paragraph(pareto_intro, styles['Normal']))
                elems.append(Spacer(1, 6))

                elems.append(PageBreak())

                # -------------------------------
                # Líneas por orden
                # -------------------------------
                lines_per_order = base.groupby('NumDoc').agg(lineas=('Articulo','nunique')).reset_index()
                dist_lines = lines_per_order.groupby('lineas').size().rename('conteo').reset_index()
                total_orders = dist_lines['conteo'].sum()
                dist_lines['%_ordenes'] = 100*dist_lines['conteo']/(total_orders if total_orders>0 else 1)
                fig2, ax2 = plt.subplots(figsize=(6,3))
                ax2.bar(dist_lines['lineas'].astype(str), dist_lines['%_ordenes'])
                ax2.set_xlabel('Líneas por orden')
                ax2.set_ylabel('% de órdenes')
                ax2.set_title('Distribución de líneas por orden')
                add_fig(fig2, 'Líneas por orden')
                
                lines_intro = """
                Este perfil muestra cuántas líneas (SKUs distintos) tiene cada pedido y qué porcentaje de órdenes corresponde a cada cantidad de líneas. 
                Permite evaluar la complejidad de los pedidos y planificar recursos de picking y personal.
                """
                elems.append(Paragraph(lines_intro, styles['Normal']))
                elems.append(Spacer(1, 6))
                elems.append(PageBreak())

                # -------------------------------
                # Cubicaje por orden
                # -------------------------------

                cubic_per_order = base.groupby('NumDoc').agg(volumen_total=('Volumen_m3','sum')).reset_index()
                vol_bins = [-1,1,2,5,10,20,50,1e9]
                vol_labels = ['≤1','1-2','2-5','5-10','10-20','20-50','>50']
                cubic_per_order['vol_bin'] = pd.cut(cubic_per_order['volumen_total'], bins=vol_bins, labels=vol_labels)
                dist_cubic = cubic_per_order.groupby('vol_bin').size().rename('conteo').reset_index()
                total_orders2 = dist_cubic['conteo'].sum()
                dist_cubic['%_ordenes'] = 100*dist_cubic['conteo']/(total_orders2 if total_orders2>0 else 1)
                fig3, ax3 = plt.subplots(figsize=(6,3))
                ax3.bar(dist_cubic['vol_bin'].astype(str), dist_cubic['%_ordenes'])
                ax3.set_xlabel('Rango volumen (pies³)')
                ax3.set_ylabel('% de órdenes')
                ax3.set_title('Distribución de volumen por orden')
                add_fig(fig3, 'Volumen por orden')

                cubic_intro = """
                El presente perfil ilustra mediante una gráfica el rango de volumen total de los pedidos y su porcentaje sobre el total de órdenes. 
                Es útil para dimensionar espacio de almacenamiento, cajas, pallets y vehículos de transporte, según requerimientos de espacio y rotación.
                """
                elems.append(Paragraph(cubic_intro, styles['Normal']))
                elems.append(Spacer(1, 6))
                elems.append(PageBreak())

                # Recalcular lv y dist_incremento para el PDF
                lv = base.groupby('NumDoc').agg(
                    lineas=('Articulo','nunique'),
                    volumen_total=('Volumen_m3','sum')
                ).reset_index()

                VOLUMEN_TARIMA = st.session_state.get('vol_tarima', 42.38)
                lv['%_carga_unidad'] = 100 * lv['volumen_total'] / VOLUMEN_TARIMA
                lv['%_carga_unidad'] = lv['%_carga_unidad'].clip(upper=100)
                carga_bins = list(range(0, 105, 5))
                carga_labels = [f'{i}-{i+5}%' for i in range(0, 100, 5)]
                lv['r_carga'] = pd.cut(lv['%_carga_unidad'], bins=carga_bins, labels=carga_labels, right=True, include_lowest=True)
                dist_incremento = lv.groupby(['r_carga']).agg(
                    pedidos=('NumDoc', 'count'),
                    lineas_prom=('lineas', 'mean')
                ).reset_index()
                dist_incremento['%_lineas_pedido'] = 100 * dist_incremento['pedidos'] / dist_incremento['pedidos'].sum()

                # Gráfica de incremento de pedidos (carga unitaria vs % líneas de pedido)
                fig_inc, ax_inc = plt.subplots(figsize=(6,3))
                ax_inc.bar(dist_incremento['r_carga'].astype(str), dist_incremento['%_lineas_pedido'])
                ax_inc.set_xlabel('% de carga unitaria (tarima)')
                ax_inc.set_ylabel('% de líneas de pedido')
                ax_inc.set_title('Distribución por incremento de pedidos')
                plt.setp(ax_inc.get_xticklabels(), rotation=60, ha='right', fontsize=7)  # Rota y reduce fuente
                add_fig(fig_inc, 'Distribución por incremento de pedidos')

                inc_intro = """
                Esta gráfica muestra la proporción de líneas de pedido según el porcentaje de carga unitaria (por ejemplo, respecto a una tarima completa).
                Permite visualizar cuántos pedidos representan cargas parciales o completas, facilitando la planificación logística y el uso eficiente de espacio.
                """
                elems.append(Paragraph(inc_intro, styles['Normal']))
                elems.append(Spacer(1, 6))
                elems.append(PageBreak())

                # -------------------------------
                # Distribución por día de la semana
                # -------------------------------

                orders_dates = base.groupby('NumDoc').agg(fecha=('Fecha','max')).reset_index()
                orders_dates['dia'] = orders_dates['fecha'].dt.day_name()
                mapping_days = {'Monday':'Lunes','Tuesday':'Martes','Wednesday':'Miércoles','Thursday':'Jueves',
                                'Friday':'Viernes','Saturday':'Sábado','Sunday':'Domingo'}
                orders_dates['dia'] = orders_dates['dia'].replace(mapping_days)
                day_order = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
                dist_days = orders_dates.groupby('dia').size().reindex(day_order).fillna(0).astype(int).rename('conteo').reset_index()
                dist_days['%_ordenes'] = 100*dist_days['conteo']/dist_days['conteo'].sum()
                fig4, ax4 = plt.subplots(figsize=(6,3))
                ax4.bar(dist_days['dia'], dist_days['%_ordenes'])
                ax4.set_xlabel('Día')
                ax4.set_ylabel('% de órdenes')
                ax4.set_title('Distribución de órdenes por día de la semana')
                add_fig(fig4, 'Órdenes por día de la semana')

                days_intro = """
                Este perfil muestra cómo se distribuyen los pedidos a lo largo de la semana y su porcentaje sobre el total. 
                Permite planificar personal, turnos y recursos logísticos en función de los picos y valles de demanda, identificando qué días presentan mayor ingreso de órdenes.
                """
                elems.append(Paragraph(days_intro, styles['Normal']))
                elems.append(PageBreak())

                # -------------------------------
                # Tabla cruzada líneas x volumen con % pedidos, Totales y Total Línea
                # -------------------------------

                lv = base.groupby('NumDoc').agg(
                    lineas=('Articulo','nunique'),
                    volumen_total=('Volumen_m3','sum')
                ).reset_index()

                # Definir rangos (misma lógica que en Streamlit)
                line_labels = ['1','2-5','6-9','10+']
                vol_labels2 = ['0-1','1-2','2-5','5-10','10-20','20+']

                # Categorizar (igual que en la app)
                lv['r_lineas'] = pd.cut(lv['lineas'], bins=[0,1,5,9,1e9], labels=line_labels, right=True, include_lowest=True)
                lv['r_vol'] = pd.cut(lv['volumen_total'], bins=[0,1,2,5,10,20,1e9], labels=vol_labels2, right=True, include_lowest=True)

                # Conteos y totales
                ct_counts = pd.crosstab(lv['r_lineas'], lv['r_vol'], dropna=False)
                ct_counts = ct_counts.reindex(index=line_labels, columns=vol_labels2, fill_value=0)
                ct_counts['Totales'] = ct_counts.sum(axis=1)

                # 🔹 Total de líneas (sumando líneas, no volumen)
                pivot_lines = pd.pivot_table(
                    lv, index='r_lineas',
                    values='lineas', aggfunc='sum', fill_value=0
                ).reindex(index=line_labels, fill_value=0)
                pivot_lines['% linea'] = (pivot_lines['lineas'] / pivot_lines['lineas'].sum() * 100).round(2)

                # Volumen total (solo para fila de "Espacio total")
                pivot_vol = pd.pivot_table(
                    lv, index='r_lineas', columns='r_vol',
                    values='volumen_total', aggfunc='sum', fill_value=0
                ).reindex(index=line_labels, columns=vol_labels2, fill_value=0).round(2)

                # Construir tabla combinada
                data_cross = []

                # Encabezado combinado
                data_cross.append(
                    ['Líneas por orden'] 
                    + ['Volumen por pedido (pies³)']*len(vol_labels2) 
                    + ['Totales','% pedidos','Total Línea','% línea']
                )
                data_cross.append(
                    [''] + vol_labels2 + ['Totales','% pedidos','Total Línea','% línea']
                )

                # Filas por r_lineas
                for idx in line_labels:
                    row_counts = ct_counts.loc[idx, vol_labels2].tolist()
                    row_total = ct_counts.loc[idx, 'Totales']
                    row_pct_pedidos = (row_total / ct_counts['Totales'].sum() * 100).round(2)
                    row_total_linea = int(pivot_lines.loc[idx, 'lineas'])  # 🔹 ahora es la suma de líneas
                    row_pct_linea = float(pivot_lines.loc[idx, '% linea'])
                    data_cross.append([idx] + row_counts + [row_total, row_pct_pedidos, row_total_linea, row_pct_linea])

                # 👉 Fila de Totales
                tot_row_counts = ct_counts[vol_labels2].sum().tolist()
                tot_total = ct_counts['Totales'].sum()
                tot_pct_pedidos = 100.0
                tot_total_linea = int(pivot_lines['lineas'].sum())  # 🔹 total líneas global
                tot_pct_linea = 100.0
                data_cross.append(['Totales'] + tot_row_counts + [tot_total, tot_pct_pedidos, tot_total_linea, tot_pct_linea])

                # Fila de % pedidos (por columna de volumen + total)
                pct_pedidos_cols = (ct_counts[vol_labels2].sum() / ct_counts['Totales'].sum() * 100).round(2).tolist()
                pct_pedidos_total = round(sum(pct_pedidos_cols), 2)
                row_pct_pedidos = ['% pedidos'] + pct_pedidos_cols + [pct_pedidos_total, '', '', '']
                data_cross.append(row_pct_pedidos)

                # Fila de volumen total por columna
                vol_values = pivot_vol[vol_labels2].sum().round(2).tolist()
                row_vol_total = ['Espacio total'] + vol_values + [pivot_vol.values.sum().round(2), '', '', '']
                data_cross.append(row_vol_total)

                # Configurar tabla PDF
                col_widths_cross = [50] + [50]*len(vol_labels2) + [50,50,50,50]
                t_cross = Table(data_cross, colWidths=col_widths_cross, hAlign='CENTER')
                t_cross.setStyle(TableStyle([
                    ('SPAN',(1,0),(len(vol_labels2),0)),  # unir fila 0 columnas de volumen
                    ('SPAN',(len(vol_labels2)+1,0),(len(vol_labels2)+1,1)),  # Totales
                    ('SPAN',(len(vol_labels2)+2,0),(len(vol_labels2)+2,1)),  # % pedidos
                    ('SPAN',(len(vol_labels2)+3,0),(len(vol_labels2)+3,1)),  # Total Línea
                    ('SPAN',(len(vol_labels2)+4,0),(len(vol_labels2)+4,1)),  # % línea
                    ('GRID',(0,0),(-1,-1),0.5,colors.black),
                    ('BACKGROUND',(0,0),(-1,1),colors.lightgrey),
                    ('BACKGROUND',(0,-3),(-1,-3),colors.lightgrey),  # Totales fila
                    ('BACKGROUND',(0,-2),(-1,-2),colors.whitesmoke),  # % pedidos
                    ('BACKGROUND',(0,-1),(-1,-1),colors.whitesmoke),  # espacio total
                    ('FONTSIZE',(0,0),(-1,-1),6),
                    ('ALIGN',(0,0),(-1,-1),'CENTER'),
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE')
                ]))
                elems.append(Paragraph('Tabla cruzada: líneas por orden vs volumen', styles['Heading2']))
                cross_intro = """
                Permite ver cuántos pedidos combinan cierta cantidad de líneas con un rango de volumen determinado, 
                junto con totales, porcentaje de pedidos y porcentaje de líneas. 
                Esto ayuda a identificar combinaciones de pedidos frecuentes o críticas y optimizar la disposición de la bodega y flujos de picking.
                """
                elems.append(Paragraph(cross_intro, styles['Normal']))
                elems.append(Spacer(1, 6))
                elems.append(t_cross)
                elems.append(Spacer(1, 10))


                # -------------------------------
                # Construir PDF
                # -------------------------------
                doc.build(elems)
                buffer.seek(0)
                st.download_button(
                    '📄 Descargar Informe PDF',
                    data=buffer.getvalue(),
                    file_name='informe_super_abc_completo.pdf',
                    mime='application/pdf'
                )
            else:
                st.info("Haz clic en el botón para generar el informe PDF.")
    else:
        st.warning("Habilita 'Generar informe PDF' en la configuración del sidebar para usar esta función.")

st.success('Cálculos finalizados. Ajusta cortes y vuelve a calcular según necesites.')

# =============================
# WMS y posiciones del almacén
# =============================
st.header('WMS y Posiciones del Almacén')

# Inicializar variables en session_state si no existen
if 'posiciones_filtradas' not in st.session_state:
    st.session_state['posiciones_filtradas'] = pd.DataFrame()
if 'registro_movimientos' not in st.session_state:
    st.session_state['registro_movimientos'] = pd.DataFrame(columns=[
        "Tipo", "Artículo", "Camas", "Cajas", "Rack Origen", "Cuerpo Origen", "Nivel Origen", "Posición Origen",
        "Rack Destino", "Cuerpo Destino", "Nivel Destino", "Posición Destino"
    ])

# =============================
# Carga del archivo de posiciones
# =============================
uploaded_positions = st.file_uploader("Cargar archivo de posiciones del almacén", type=["xlsx"])
if uploaded_positions:
    xls = pd.ExcelFile(uploaded_positions)
    hojas = xls.sheet_names
    hoja_seleccionada = st.selectbox("Selecciona la hoja a cargar", hojas)
    
    if hoja_seleccionada:
        columnas_necesarias = ['Rack', 'Cuerpo', 'Nivel', 'Posición', 'Zona', 'Artículo', 'Camas', 'Cajas']
        posiciones = pd.read_excel(uploaded_positions, sheet_name=hoja_seleccionada, usecols=columnas_necesarias)
        st.success(f"Hoja '{hoja_seleccionada}' cargada correctamente")

        # Limpieza básica
        posiciones['Artículo'] = posiciones['Artículo'].astype(str).str.strip().str.upper()
        posiciones['Nivel'] = pd.to_numeric(posiciones['Nivel'], errors='coerce').fillna(0).astype(int)
        posiciones['Posición'] = posiciones['Posición'].astype(str).fillna('')
        
        # Separar múltiples SKUs por fila
        posiciones['Artículos_lista'] = posiciones['Artículo'].str.split(' - ')

        # Filtrar racks y niveles relevantes
        posiciones_filtradas = posiciones[
            ((posiciones['Rack'] == 1) & (posiciones['Nivel'].isin([1,2,3,4,5]))) |
            ((posiciones['Rack'] == 2) & (posiciones['Nivel'].isin([1,2,3,4,5])))
        ].copy()
        
        st.session_state['posiciones_filtradas'] = posiciones_filtradas
        st.write("Posiciones del almacén:")
        st.dataframe(posiciones_filtradas)

# =============================
# Verificar correspondencia con Súper ABC
# =============================
if 'by_item' in st.session_state and not st.session_state['posiciones_filtradas'].empty:
    by_item = st.session_state['by_item']
    posiciones_filtradas = st.session_state['posiciones_filtradas']

    articulos_posiciones = set([sku for sublist in posiciones_filtradas['Artículos_lista'] for sku in sublist])
    articulos_by_item = set(by_item.index)
    articulos_faltantes = articulos_posiciones - articulos_by_item - {"", "VACÍO"}
    if articulos_faltantes:
        st.warning(f"⚠️ Los siguientes artículos en posiciones no están en el Súper ABC: {articulos_faltantes}")
    else:
        st.success("✅ Todos los artículos en posiciones tienen correspondencia en el Súper ABC.")

# =============================
# Función para determinar zona por SKU
# =============================
zonas_prioridad = {'Oro': 3, 'Plata': 2, 'Bronce': 1, 'MP': 0}

def zona_ideal_sku(sku):
    sku = sku.strip().upper()
    if sku in ["", "VACÍO"]:
        return None  # vacíos no se consideran
    if sku in by_item.index:
        zona = by_item.loc[sku, 'Zona_Bodega']
        return zona if pd.notna(zona) else 'Bronce'
    return 'MP'

def zona_prioritaria_fila(articulos):
    """
    Determina la zona prioritaria de una fila (tarima) con múltiples SKUs,
    ignorando vacíos. Retorna None si todos son vacíos.
    """
    skus = [sku.strip().upper() for sku in articulos if sku.strip().upper() not in ["", "VACÍO"]]
    if not skus:
        return None
    zonas = [zona_ideal_sku(sku) for sku in skus if zona_ideal_sku(sku) is not None]
    if not zonas:
        return None
    zonas_sorted = sorted(zonas, key=lambda z: zonas_prioridad.get(z,0), reverse=True)
    return zonas_sorted[0]

# =============================
# Reubicación automática de tarimas
# =============================
if not st.session_state['posiciones_filtradas'].empty:
    posiciones_filtradas = st.session_state['posiciones_filtradas'].copy()
    posiciones_filtradas['Zona_Ideal_Fila'] = posiciones_filtradas['Artículos_lista'].apply(zona_prioritaria_fila)

    # Filas mal ubicadas (ignorando vacíos)
    mal_ubicadas = posiciones_filtradas[
        (posiciones_filtradas['Zona'] != posiciones_filtradas['Zona_Ideal_Fila']) &
        (posiciones_filtradas['Zona_Ideal_Fila'].notna())
    ].copy()

    if not mal_ubicadas.empty:
        mal_ubicadas['Ubicación_Actual'] = mal_ubicadas.apply(
            lambda r: f"Rack {r['Rack']}, Cuerpo {r['Cuerpo']}, Nivel {r['Nivel']}, Posición {r['Posición']}", axis=1
        )
        st.subheader("📊 Filas mal ubicadas")
        st.dataframe(mal_ubicadas[['Artículos_lista','Zona','Zona_Ideal_Fila','Ubicación_Actual']])
        # -----------------------------
        # Gráfica de filas mal ubicadas por zona (excluyendo MP)
        # -----------------------------
        # Filtrar MP y contar por la zona actual
        conteo_zonas = mal_ubicadas[mal_ubicadas['Zona'] != 'MP']['Zona'].value_counts().reindex(['Oro','Plata','Bronce'], fill_value=0).reset_index()
        conteo_zonas.columns = ['Zona', 'Filas Mal Ubicadas']

        # Colores fijos
        colores = {'Oro': '#FFD700', 'Plata': '#C0C0C0', 'Bronce': '#CD7F32'}

        fig = px.bar(
            conteo_zonas,
            x='Zona',
            y='Filas Mal Ubicadas',
            color='Zona',
            color_discrete_map=colores,
            text='Filas Mal Ubicadas',
            title="Filas mal ubicadas por Zona (Oro, Plata, Bronce)"
        )
        fig.update_layout(showlegend=False)
        st.plotly_chart(fig)

        posiciones_reconfiguradas = posiciones_filtradas.copy()

        for idx, fila in mal_ubicadas.iterrows():
            zona_objetivo = fila['Zona_Ideal_Fila']

            # Buscar slots vacíos en la zona objetivo (priorizar niveles bajos y cuerpos cercanos)
            vacios = posiciones_reconfiguradas[
                (posiciones_reconfiguradas['Artículo'].str.upper().isin(["", "VACÍO"])) &
                (posiciones_reconfiguradas['Zona'] == zona_objetivo)
            ].sort_values(by=['Nivel', 'Cuerpo', 'Rack', 'Posición'])

            if not vacios.empty:
                slot = vacios.iloc[0]
                # Mover tarima completa
                posiciones_reconfiguradas.loc[slot.name, ['Artículo','Camas','Cajas']] = \
                    [fila['Artículo'], fila['Camas'], fila['Cajas']]
                posiciones_reconfiguradas.loc[idx, ['Artículo','Camas','Cajas']] = ["VACÍO", np.nan, np.nan]
            else:
                # fallback a zonas de menor prioridad (también priorizando niveles bajos)
                prioridad_actual = zonas_prioridad[zona_objetivo]
                for z, p in sorted(zonas_prioridad.items(), key=lambda x: -x[1]):
                    if p < prioridad_actual:
                        vacios_alt = posiciones_reconfiguradas[
                            (posiciones_reconfiguradas['Artículo'].str.upper().isin(["", "VACÍO"])) &
                            (posiciones_reconfiguradas['Zona'] == z)
                        ].sort_values(by=['Nivel', 'Cuerpo', 'Rack', 'Posición'])
                        if not vacios_alt.empty:
                            slot = vacios_alt.iloc[0]
                            posiciones_reconfiguradas.loc[slot.name, ['Artículo','Camas','Cajas']] = \
                                [fila['Artículo'], fila['Camas'], fila['Cajas']]
                            posiciones_reconfiguradas.loc[idx, ['Artículo','Camas','Cajas']] = ["VACÍO", np.nan, np.nan]
                            break

        st.subheader("🔄 Reubicación recomendada")
        st.dataframe(posiciones_reconfiguradas)

        # =============================
        # Sección para descarga de resultados
        # =============================
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            posiciones_reconfiguradas.to_excel(writer, index=False, sheet_name='Posiciones')

        output.seek(0)  # mover el cursor al inicio del archivo

        st.download_button(
            label="📥 Descargar posiciones reconfiguradas",
            data=output,
            file_name='posiciones_reconfiguradas.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


    import pandas as pd
    from openpyxl import load_workbook

    def guardar_configuracion_y_movimientos(posiciones, movimientos, archivo_salida="almacen_actualizado.xlsx"):
        """
        Guarda la configuración actual del almacén y los movimientos realizados en un archivo Excel.

        Args:
            posiciones (pd.DataFrame): DataFrame con la configuración actual del almacén.
            movimientos (pd.DataFrame): DataFrame con el registro de movimientos realizados.
            archivo_salida (str): Nombre del archivo Excel de salida.

        Returns:
            str: Ruta del archivo Excel generado.
        """
        with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
            # Guardar la configuración actual del almacén
            posiciones.to_excel(writer, sheet_name="Configuración Actual", index=False)

            # Guardar el registro de movimientos
            movimientos.to_excel(writer, sheet_name="Registro de Movimientos", index=False)

        return archivo_salida

    # Registro de movimientos
    st.subheader("📋 Registro de Movimientos")

    # Selección del tipo de movimiento
    tipo_movimiento = st.selectbox("Tipo de movimiento", ["Ingreso", "Salida", "Traslado"])

    # Datos comunes para todos los movimientos
    articulo = st.text_input("Código del artículo")
    camas = st.number_input("Camas", min_value=0, step=1)
    cajas = st.number_input("Cajas", min_value=0, step=1)

    # Crear un DataFrame para registrar los movimientos
    if "registro_movimientos" not in st.session_state:
        st.session_state["registro_movimientos"] = pd.DataFrame(columns=[
            "Tipo", "Artículo", "Camas", "Cajas", "Rack Origen", "Cuerpo Origen", "Nivel Origen", "Posición Origen",
            "Rack Destino", "Cuerpo Destino", "Nivel Destino", "Posición Destino"
        ])

    registro_movimientos = st.session_state["registro_movimientos"]

    if tipo_movimiento == "Ingreso":
        # Mostrar ubicaciones sugeridas antes de registrar el ingreso
        if articulo in by_item.index:
            prioridad_actual = by_item.loc[articulo, 'Clase_SuperABC']
            zona_prioridad = map_zone(prioridad_actual)

            # Filtrar posiciones disponibles según la zona de prioridad
            disponibles = posiciones_filtradas[posiciones_filtradas['Artículo'].isna()]
            if zona_prioridad == 'Oro':
                sugerencias = disponibles[
                    (disponibles['Rack'] == 1) &
                    (disponibles['Cuerpo'].isin([1, 2])) &
                    (disponibles['Nivel'].isin([1, 2]))
                ]
            elif zona_prioridad == 'Plata':
                sugerencias = disponibles[
                    (disponibles['Cuerpo'].isin([3, 4, 5])) &
                    (disponibles['Nivel'].isin([1, 2]))
                ]
            else:  # Bronce
                sugerencias = disponibles[
                    (disponibles['Nivel'] == 3)
                ]

            if sugerencias.empty:
                st.warning(f"No se encontraron ubicaciones disponibles para el artículo '{articulo}' en la zona {zona_prioridad}.")
            else:
                st.write(f"📍 Ubicaciones sugeridas para el artículo '{articulo}' (Zona: {zona_prioridad}):")
                st.dataframe(sugerencias.head(5))  # Mostrar las primeras 5 sugerencias

        # Datos específicos para ingreso
        rack = st.number_input("Rack (destino)", min_value=1, step=1)
        cuerpo = st.number_input("Cuerpo (destino)", min_value=1, step=1)
        nivel = st.number_input("Nivel (destino)", min_value=1, max_value=5, step=1)
        posicion = st.text_input("Posición (destino, ej. I1, C1, D1)")

        if st.button("Registrar Ingreso"):
            # Verificar si la ubicación está disponible
            ubicacion_disponible = posiciones_filtradas[
                (posiciones_filtradas['Rack'] == rack) &
                (posiciones_filtradas['Cuerpo'] == cuerpo) &
                (posiciones_filtradas['Nivel'] == nivel) &
                (posiciones_filtradas['Posición'] == posicion) &
                (posiciones_filtradas['Artículo'].isna())
            ]

            if not ubicacion_disponible.empty:
                st.success(f"Ingreso registrado: {articulo} - Rack {rack}, Cuerpo {cuerpo}, Nivel {nivel}, Posición {posicion}, Camas {camas}, Cajas {cajas}")
                # Actualizar la ubicación en el DataFrame
                posiciones_filtradas.loc[
                    (posiciones_filtradas['Rack'] == rack) &
                    (posiciones_filtradas['Cuerpo'] == cuerpo) &
                    (posiciones_filtradas['Nivel'] == nivel) &
                    (posiciones_filtradas['Posición'] == posicion),
                    ['Artículo', 'Camas', 'Cajas']
                ] = [articulo, camas, cajas]

                # Registrar el movimiento
                nuevo_movimiento = pd.DataFrame([{
                    "Tipo": "Ingreso",
                    "Artículo": articulo,
                    "Camas": camas,
                    "Cajas": cajas,
                    "Rack Origen": None,
                    "Cuerpo Origen": None,
                    "Nivel Origen": None,
                    "Posición Origen": None,
                    "Rack Destino": rack,
                    "Cuerpo Destino": cuerpo,
                    "Nivel Destino": nivel,
                    "Posición Destino": posicion
                }])

                registro_movimientos = pd.concat([registro_movimientos, nuevo_movimiento], ignore_index=True)
                st.session_state["registro_movimientos"] = registro_movimientos

                # Guardar la configuración y los movimientos en un archivo Excel
                archivo_actualizado = guardar_configuracion_y_movimientos(posiciones_filtradas, registro_movimientos)
                st.success("Archivo actualizado con el movimiento registrado.")
                with open(archivo_actualizado, "rb") as file:
                    st.download_button(
                        label="📥 Descargar archivo actualizado",
                        data=file,
                        file_name=archivo_actualizado,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.error("La ubicación seleccionada no está disponible. Por favor, selecciona una de las ubicaciones sugeridas.")

    import json

    def guardar_estado_json():
        """
        Guarda el estado completo de la aplicación en un archivo JSON.
        """
        estado = {
            "posiciones_filtradas": posiciones_filtradas.to_dict() if 'posiciones_filtradas' in locals() else None,
            "registro_movimientos": registro_movimientos.to_dict() if 'registro_movimientos' in locals() else None,
            "criterios_seleccionados": st.session_state.get("criterios_seleccionados", []),
            "cortes_abc": st.session_state.get("cortes_abc", {}),
            "by_item": st.session_state.get("by_item", {}).to_dict() if "by_item" in st.session_state else None,
            "ubicaciones_clasificacion": st.session_state.get("ubicaciones_clasificacion", {}).to_dict() if "ubicaciones_clasificacion" in st.session_state else None
        }
        with open("estado_app.json", "w") as f:
            json.dump(estado, f)
        st.success("Estado guardado correctamente.")

    def cargar_estado_json():
        """
        Carga el estado completo de la aplicación desde un archivo JSON.
        """
        global posiciones_filtradas, registro_movimientos
        try:
            with open("estado_app.json", "r") as f:
                estado = json.load(f)
            
            # Restaurar los datos en las variables globales y session_state
            if estado.get("posiciones_filtradas"):
                posiciones_filtradas = pd.DataFrame.from_dict(estado["posiciones_filtradas"])
            if estado.get("registro_movimientos"):
                registro_movimientos = pd.DataFrame.from_dict(estado["registro_movimientos"])
            st.session_state["criterios_seleccionados"] = estado.get("criterios_seleccionados", [])
            st.session_state["cortes_abc"] = estado.get("cortes_abc", {})
            if estado.get("by_item"):
                st.session_state["by_item"] = pd.DataFrame.from_dict(estado["by_item"])
            if estado.get("ubicaciones_clasificacion"):
                st.session_state["ubicaciones_clasificacion"] = pd.DataFrame.from_dict(estado["ubicaciones_clasificacion"])
            
            st.success("Estado cargado correctamente.")
        except FileNotFoundError:
            st.warning("No se encontró un archivo de estado previo. Inicia desde cero.")
        except Exception as e:
            st.error(f"Error al cargar el estado: {e}")

if st.button("Cargar estado previo"):
    cargar_estado_json()

if st.button("Guardar estado actual"):
    guardar_estado_json()
